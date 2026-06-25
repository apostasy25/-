# -*- coding: utf-8 -*-
"""
M3 정산서 시트 생성기 (사업자 세금계산서 양식).

실제 발송 양식과 동일 구조로 openpyxl 작성:
 · 상단 증빙표 : 증빙서류·작성일자·품목명·공급가액(SUMIFS)·부가세(ROUND)·실지급(G+H)·합계
 · 작품 섹션   : 순번·국가·구분·플랫폼·작품·런칭일·서비스월·총매출·정산기준순매출·RS·정산금액(ROUND(J*K))·비고
 · MG 작품     : 전체이력 상세 + MG총액행 + 잔여 MG행, 증빙표 실지급 = IF(잔여>0,잔여,"MG 차감중")
값은 RAW(매출 연재 + 기타수익 광고)에서 재구성(m3_builder), 직전 정산서 미사용.
"""
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import m3_builder as B

_TH = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)
_HDRFILL = PatternFill("solid", fgColor="F2F2F2")
_PROOFFILL = PatternFill("solid", fgColor="FF9900")    # 증빙표 헤더(주황) — 원본 일치
_DETAILFILL = PatternFill("solid", fgColor="E7E6E6")   # 연재상세 헤더(회색) — 원본 일치
_AMTFILL = PatternFill("solid", fgColor="FFF2CC")      # 공급가액(연노랑) — 원본 일치
_BOLD = Font(name="맑은 고딕", size=10, bold=True)
_NORM = Font(name="맑은 고딕", size=10)
_CEN = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

DET_HDR = ["순번", "국가", "구분", "플랫폼명", "작품명", "런칭일", "서비스월",
           "총매출", "정산기준순매출", "원작사 RS율", "정산금액", "비고"]


def _wcell(ws, r, c, v, font=_NORM, align=_LEFT, border=True, fmt=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    cell.alignment = align
    if border:
        cell.border = _BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def _svc_ym(d):
    """행의 서비스월을 (연,월) 키로. 이월('-')은 맨 앞, 파싱 불가는 맨 뒤."""
    import re as _re
    sv = d.get("서비스월")
    if sv in ("-", None, ""):
        return (0, 0)
    if hasattr(sv, "year"):
        return (sv.year, sv.month)
    m = _re.search(r"(\d{4})\D*(\d{1,2})", str(sv))
    return (int(m.group(1)), int(m.group(2))) if m else (9999, 99)


def _interleave(a, b):
    """연재(a)·광고(b)를 서비스월 기준으로 합쳐 월별로 모이게 한다(누적 MG 작품용).
    같은 달이면 광고수익 먼저, 독점(연재)을 마지막에 둔다 — 전월 정산서에 당월분을 덧붙인 형태와 일치.
    안정 정렬로 같은 키 내 원래 순서 보존."""
    tagged = [(_svc_ym(d), 1, i, d) for i, d in enumerate(a)] \
        + [(_svc_ym(d), 0, i, d) for i, d in enumerate(b)]
    tagged.sort(key=lambda t: (t[0], t[1], t[2]))
    return [t[3] for t in tagged]


def collect_vendor_ctx(vendor_name, 정산서월, works, rsmap, mgset,
                       sales_path, etc_path, mg_master, sales_cache=None, etc_cache=None,
                       rs_override=None, period_months=None, 품목_label=None, 이월_map=None,
                       발송이메일="정산팀\nterapin_toonbill@terapinstudios.co.kr",
                       마감일=None, 지급일=None, 직접정산=False, 작성일=None,
                       증빙구분="세금계산서", type_key="사업자"):
    """build_vendor_sheet와 동일 로직으로 데이터만 수집 → template_fill.fill_business용 ctx 반환.
       MG 작품은 당월 발생분(이번 지급월 지급시트 / 광고 원작사정산월=정산서월)에 당월=True 플래그."""
    y, mo = 정산서월.split("-")
    이월_map = 이월_map or {}
    sheet = B._ymd_to_sheet(B._next_month(정산서월))      # 이번 정산 지급월 시트
    import datetime as _dt

    def _is_now(d):                                        # MG 당월(신규 발생분) 판정
        if d.get("지급시트") == sheet:
            return True
        if d.get("구분") == "광고수익" and d.get("서비스월") == 정산서월:
            return True
        return False

    works_ctx = []
    for w in works:
        rs = rsmap.get(w, 0.1)
        rs_연재 = (rs_override or {}).get(w, rs)
        is_mg = w in mgset
        mg_total = 0
        held = False
        if period_months:
            연재 = B.연재상세(sales_path, w, rs_연재, 서비스월_set=period_months, cache=sales_cache)
            광고 = B.광고상세(etc_path, w, rs, 원작사정산월_cutoff=None,
                          원작사정산월_set=period_months, cache=etc_cache)
        elif is_mg:
            mg_total = _mg_total(mg_master, w)
            연재 = B.연재상세(sales_path, w, rs_연재, 서비스월_cutoff=_prev_month(정산서월), cache=sales_cache)
            광고 = B.광고상세(etc_path, w, rs, 원작사정산월_cutoff=정산서월, cache=etc_cache)
        else:
            iw = 이월_map.get(w)
            연재 = B.연재상세(sales_path, w, rs_연재, 당월_지급시트=sheet, cache=sales_cache)
            광고 = B.광고상세(etc_path, w, rs, 원작사정산월_cutoff=None, 당월only=정산서월, cache=etc_cache)
            if iw and str(iw.get("증빙상태")) == "수취" and iw.get("합산월") == 정산서월:
                이월순 = round(iw.get("이월순매출") or 0)
                연재 = [{"국가": "-", "구분": "-", "플랫폼": "-", "런칭": "-", "서비스월": "-",
                       "총매출": "-", "정산기준순매출": 이월순, "RS": rs,
                       "정산금액": round(이월순 * rs), "비고": "이월 정산금"}] + 연재
            elif iw and str(iw.get("증빙상태")) == "미수취":
                held = True
        # 일반 작품: 독점 전체 ▶ 광고수익 전체. 누적 MG 작품만 월별(독점▶광고) 나열.
        rows = _interleave(연재, 광고) if is_mg else (연재 + 광고)
        if not rows and not is_mg:
            continue
        ctx_rows = []
        for d in rows:
            ctx_rows.append({
                "국가": d.get("국가"), "구분": d.get("구분"), "플랫폼": d.get("플랫폼"),
                "작품": w, "런칭": d.get("런칭"), "서비스월": d.get("서비스월"),
                "총매출": d.get("총매출"), "정산기준순매출": d.get("정산기준순매출"),
                "RS": d.get("RS"), "비고": d.get("비고") or "",
                "당월": _is_now(d) if is_mg else False,
            })
        wctx = {"작품": w, "rows": ctx_rows, "held": held}
        if is_mg:
            wctx["mg_total"] = mg_total
        works_ctx.append(wctx)

    작성 = "-" if 직접정산 else (작성일 or _last_day(int(y), int(mo)))
    _ny, _nm = (int(y) + 1, 1) if int(mo) == 12 else (int(y), int(mo) + 1)
    마감 = "-" if 직접정산 else (마감일 or _dt.datetime(_ny, _nm, 10))
    지급 = "-" if 직접정산 else (지급일 or _dt.datetime(_ny, _nm, 15))
    return {
        "year": int(y), "month": int(mo), "vendor": vendor_name, "type_key": type_key,
        "증빙구분": 증빙구분, "작성일자": 작성, "마감": 마감, "지급": 지급,
        "email": 발송이메일, "직접정산": 직접정산,
        "품목_label": 품목_label, "works": works_ctx,
    }


def build_vendor_sheet(ws, vendor_name, 정산서월, works, rsmap, mgset,
                       sales_path, etc_path, mg_master, sales_cache=None, etc_cache=None,
                       rs_override=None, period_months=None, 품목_label=None, 이월_map=None,
                       발송이메일="정산팀\nterapin_toonbill@terapinstudios.co.kr",
                       마감일=None, 지급일=None, 직접정산=False, 작성일=None):
    """ws에 한 업체 정산서를 그린다. works=표준작품명 리스트(증빙표 순서).
       rs_override: {작품: 연재RS} — 누적구간 tier.
       period_months: 분기/반기 서비스월 집합(누적). 품목_label: 품목명 기간 표기.
       이월_map: {작품: {증빙상태, 발생월, 합산월}} — 증빙 미수취 보류/수취 합산."""
    y, mo = 정산서월.split("-")
    이월_map = 이월_map or {}
    이월info = {}                          # {작품: {합산발생월, held}}
    # 제목
    ws.cell(2, 2, f"{y}년 {mo}월 {vendor_name} 원작료 정산서").font = Font(
        name="맑은 고딕", size=14, bold=True)

    # ── 1단계: 작품별 상세 수집 후 빈 작품(비MG) 제외 ──
    sheet = B._ymd_to_sheet(B._next_month(정산서월))
    items = []                            # (작품, rs, is_mg, mg_total, rows)
    for w in works:
        rs = rsmap.get(w, 0.1)
        rs_연재 = (rs_override or {}).get(w, rs)   # 누적구간 tier RS(만년 0.20 등)
        is_mg = w in mgset
        mg_total = 0
        if period_months:                  # 분기/반기 누적(서비스월 ∈ 기간, 정규 배치 1곳)
            연재 = B.연재상세(sales_path, w, rs_연재, 서비스월_set=period_months, cache=sales_cache)
            광고 = B.광고상세(etc_path, w, rs, 원작사정산월_cutoff=None,
                          원작사정산월_set=period_months, cache=etc_cache)
        elif is_mg:
            mg_total = _mg_total(mg_master, w)
            연재 = B.연재상세(sales_path, w, rs_연재, 서비스월_cutoff=_prev_month(정산서월), cache=sales_cache)
            광고 = B.광고상세(etc_path, w, rs, 원작사정산월_cutoff=정산서월, cache=etc_cache)
        else:
            iw = 이월_map.get(w)
            연재 = B.연재상세(sales_path, w, rs_연재, 당월_지급시트=sheet, cache=sales_cache)
            광고 = B.광고상세(etc_path, w, rs, 원작사정산월_cutoff=None, 당월only=정산서월, cache=etc_cache)
            if iw and str(iw.get("증빙상태")) == "수취" and iw.get("합산월") == 정산서월:
                # 합산: 직전 정산서 정산기준순매출 합을 '이월 정산금' 단일 행으로 선두 추가
                이월순 = round(iw.get("이월순매출") or 0)
                이월행 = {"국가": "-", "구분": "-", "플랫폼": "-", "런칭": "-",
                        "서비스월": "-", "총매출": "-", "정산기준순매출": 이월순,
                        "RS": rs, "정산금액": round(이월순 * rs), "비고": "이월 정산금"}
                연재 = [이월행] + 연재
                이월info[w] = {"합산발생월": iw.get("발생월"), "held": False}
            elif iw and str(iw.get("증빙상태")) == "미수취":
                # 보류: 당월분 계산하되 증빙 미수취로 지급 보류(누적)
                이월info[w] = {"합산발생월": None, "held": True}
            else:
                연재 = B.연재상세(sales_path, w, rs_연재, 당월_지급시트=sheet, cache=sales_cache)
                광고 = B.광고상세(etc_path, w, rs, 원작사정산월_cutoff=None, 당월only=정산서월, cache=etc_cache)
        rows = 연재 + 광고
        if not rows and not is_mg:        # 당월 데이터 없는 비MG 작품 제외
            continue
        items.append((w, rs, is_mg, mg_total, rows))

    # ── 증빙표 헤더(행5) ──
    proof_hdr = {2: "증빙서류", 4: "작성일자", 5: "품목명",
                 7: "공급가액", 8: "부가세", 9: "실지급액",
                 10: "(세금)계산서 발송 이메일 주소", 13: "마감일", 14: "지급일"}
    for c, t in proof_hdr.items():
        _wcell(ws, 5, c, t, _BOLD, _CEN, fmt=None).fill = _PROOFFILL

    n = len(items)
    proof_first = 6
    total_row = proof_first + n
    sec_start = total_row + 2

    # ── 2단계: 섹션 그리기 ──
    sec_pos = {}
    cur = sec_start
    for i, (w, rs, is_mg, mg_total, rows) in enumerate(items, 1):
        # 섹션 헤더
        _wcell(ws, cur, 2, f"{i}. 웹툰 <{w}> 원작료 정산 상세내역", _BOLD, _LEFT, border=False)
        hdr_r = cur + 1
        for c, t in enumerate(DET_HDR, 2):
            _wcell(ws, hdr_r, c, t, _BOLD, _CEN).fill = _DETAILFILL
        first = hdr_r + 1
        rr = first
        for idx, d in enumerate(rows, 1):
            _wcell(ws, rr, 2, idx, _NORM, _CEN)
            _wcell(ws, rr, 3, d["국가"], _NORM, _CEN)
            _wcell(ws, rr, 4, d["구분"], _NORM, _CEN)
            _wcell(ws, rr, 5, d["플랫폼"], _NORM, _LEFT)
            _wcell(ws, rr, 6, w, _NORM, _LEFT)
            _wcell(ws, rr, 7, d.get("런칭") or "-", _NORM, _CEN)
            _wcell(ws, rr, 8, d["서비스월"], _NORM, _CEN)
            _wcell(ws, rr, 9, d["총매출"], _NORM, _RIGHT, fmt="#,##0")
            _wcell(ws, rr, 10, d["정산기준순매출"], _NORM, _RIGHT, fmt="#,##0")
            _wcell(ws, rr, 11, d["RS"], _NORM, _CEN, fmt="0.0%")   # 행별 RS(연재 tier / 광고 평면)
            _wcell(ws, rr, 12, f"=ROUND(J{rr}*K{rr},0)", _NORM, _RIGHT, fmt="#,##0")
            _wcell(ws, rr, 13, d.get("비고") or "", _NORM, _LEFT)
            rr += 1
        last = rr - 1
        sumrow = None
        mg잔여 = None
        if is_mg:
            # MG총액 행
            _wcell(ws, rr, 5, "선인세(MG) 총액", _NORM, _LEFT)
            _wcell(ws, rr, 12, mg_total, _NORM, _RIGHT, fmt="#,##0")
            mgrow = rr
            rr += 1
            # 잔여 MG 행: L = SUM(상세 L) - MG총액 ; M = "MG 차감중"
            _wcell(ws, rr, 2, "잔여 MG", _BOLD, _LEFT)
            _wcell(ws, rr, 12, f"=SUM(L{first}:L{last})-L{mgrow}", _BOLD, _RIGHT, fmt="#,##0")
            _wcell(ws, rr, 13, "MG 차감중", _BOLD, _CEN)
            mg잔여 = rr
            rr += 1
        else:
            _wcell(ws, rr, 2, "합계", _BOLD, _LEFT)
            _wcell(ws, rr, 12, f"=SUM(L{first}:L{last})", _BOLD, _RIGHT, fmt="#,##0")
            sumrow = rr
            rr += 1
        sec_pos[w] = dict(first=first, last=last, sumrow=sumrow, mg잔여=mg잔여, is_mg=is_mg)
        cur = rr + 1                      # 섹션 간 여백 1행

    # ── 증빙표 본문(행6~) ──
    import datetime
    작성 = "-" if 직접정산 else (작성일 or _last_day(int(y), int(mo)))
    ws.cell(proof_first, 2, "세금계산서").font = _NORM        # 증빙서류(병합 영역 대표)
    ws.cell(proof_first, 2).alignment = _CEN
    ws.cell(proof_first, 4, 작성).font = _NORM
    ws.cell(proof_first, 4).alignment = _CEN
    if not 직접정산:
        ws.cell(proof_first, 4).number_format = "yyyy-mm-dd"
    # 발송이메일(첫 행) + 마감일·지급일(업체별) / 카카오 등 직접정산은 "-"
    import datetime as _dt
    _ny, _nm = (int(y) + 1, 1) if int(mo) == 12 else (int(y), int(mo) + 1)
    _마감 = 마감일 or _dt.datetime(_ny, _nm, 10)
    _지급 = 지급일 or _dt.datetime(_ny, _nm, 15)
    _wcell(ws, proof_first, 10, "-" if 직접정산 else 발송이메일, _NORM, _LEFT)
    if 직접정산:
        _wcell(ws, proof_first, 13, "-", _NORM, _CEN)
        _wcell(ws, proof_first, 14, "-", _NORM, _CEN)
    else:
        _wcell(ws, proof_first, 13, _마감, _NORM, _CEN, fmt="yyyy-mm-dd")
        _wcell(ws, proof_first, 14, _지급, _NORM, _CEN, fmt="yyyy-mm-dd")
    for i, (w, rs, is_mg, mg_total, rows) in enumerate(items):
        r = proof_first + i
        info = 이월info.get(w, {})
        품목 = f"원작료_{품목_label or (mo + '월')} 정산 <{w}>"
        _wcell(ws, r, 5, 품목, _NORM, _LEFT)
        p = sec_pos[w]
        if p["is_mg"]:
            _wcell(ws, r, 7, 0, _NORM, _RIGHT, fmt="#,##0").fill = _AMTFILL
            _wcell(ws, r, 8, 0, _NORM, _RIGHT, fmt="#,##0")
            jan = p["mg잔여"]
            _wcell(ws, r, 9, f'=IF(L{jan}>0,L{jan},M{jan})', _NORM, _RIGHT)
        elif info.get("held"):
            # 증빙 미수취 → 공급가액·부가세는 표시, 실지급은 보류(합계 제외)
            _wcell(ws, r, 7, f"=SUMIFS($L:$L,$F:$F,$F{p['first']})", _NORM, _RIGHT, fmt="#,##0").fill = _AMTFILL
            _wcell(ws, r, 8, f"=ROUND(G{r}*10%,0)", _NORM, _RIGHT, fmt="#,##0")
            _wcell(ws, r, 9, "증빙 미수취-보류", _NORM, _CEN)
        else:
            _wcell(ws, r, 7, f"=SUMIFS($L:$L,$F:$F,$F{p['first']})", _NORM, _RIGHT, fmt="#,##0").fill = _AMTFILL
            _wcell(ws, r, 8, f"=ROUND(G{r}*10%,0)", _NORM, _RIGHT, fmt="#,##0")
            _wcell(ws, r, 9, f"=G{r}+H{r}", _NORM, _RIGHT, fmt="#,##0")
    # 합계행
    _wcell(ws, total_row, 2, "합계", _BOLD, _CEN)
    _wcell(ws, total_row, 7, f"=SUM(G{proof_first}:G{total_row-1})", _BOLD, _RIGHT, fmt="#,##0")
    _wcell(ws, total_row, 8, f"=SUM(H{proof_first}:H{total_row-1})", _BOLD, _RIGHT, fmt="#,##0")
    _wcell(ws, total_row, 9, f"=SUM(I{proof_first}:I{total_row-1})", _BOLD, _RIGHT, fmt="#,##0")
    if not 직접정산:
        _wcell(ws, total_row, 10,
               "(세금)계산서는 1장으로 발행해 주시되 품목 나눠 기재 부탁드립니다.", _NORM, _LEFT)

    # 열 너비
    widths = {2: 6, 3: 9, 4: 12, 5: 26, 6: 22, 7: 13, 8: 13, 9: 13,
              10: 14, 11: 9, 12: 12, 13: 16}
    for c, wd in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = wd
    return sec_pos


def _mg_total(mg_master, work):
    from openpyxl import load_workbook
    ws = load_workbook(mg_master, data_only=True)["선인세MG마스터"]
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 2).value) == work:
            try:
                return float(ws.cell(r, 3).value)
            except (TypeError, ValueError):
                return 0
    return 0


def _last_day(y, mo):
    import calendar
    import datetime
    return datetime.datetime(y, mo, calendar.monthrange(y, mo)[1])


def _prev_month(ym):
    y, m = map(int, ym.split("-"))
    return f"{y - 1}-12" if m == 1 else f"{y}-{m - 1:02d}"


# ── 개인(작가) 정산서: 원천징수 3.3%, 부가세 없음, 웹툰+기타 2섹션 ──
P_HDR = ["순번", "국가", "구분", "플랫폼명", "작품명", "런칭일", "서비스월",
         "정산기준순매출", "원작 RS율", "정산금액", "비고"]


def _detail_block(ws, start_row, title, hdr5, rows, namecol_label):
    """개인용 상세 블록 그리기 → (합계행, 다음행)."""
    _wcell(ws, start_row, 2, title, _BOLD, _LEFT, border=False)
    hr = start_row + 1
    hdr = list(P_HDR)
    hdr[4] = namecol_label                      # 작품명 또는 광고명
    for c, t in enumerate(hdr, 2):
        _wcell(ws, hr, c, t, _BOLD, _CEN).fill = _HDRFILL
    first = hr + 1
    r = first
    for i, d in enumerate(rows, 1):
        _wcell(ws, r, 2, i, _NORM, _CEN)
        _wcell(ws, r, 3, d["국가"], _NORM, _CEN)
        _wcell(ws, r, 4, d["구분"], _NORM, _CEN)
        _wcell(ws, r, 5, d["플랫폼"], _NORM, _LEFT)
        _wcell(ws, r, 6, d.get("작품", ""), _NORM, _LEFT)
        _wcell(ws, r, 7, d.get("런칭") or "-", _NORM, _CEN)
        _wcell(ws, r, 8, d["서비스월"], _NORM, _CEN)
        _wcell(ws, r, 9, d["정산기준순매출"], _NORM, _RIGHT, fmt="#,##0")
        _wcell(ws, r, 10, d["RS"], _NORM, _CEN, fmt="0.0%")
        _wcell(ws, r, 11, f"=ROUND(I{r}*J{r},0)", _NORM, _RIGHT, fmt="#,##0")
        _wcell(ws, r, 12, d.get("비고") or "", _NORM, _LEFT)
        r += 1
    last = r - 1
    _wcell(ws, r, 2, "합계", _BOLD, _CEN)
    val = f"=SUM(K{first}:K{last})" if last >= first else 0
    _wcell(ws, r, 11, val, _BOLD, _RIGHT, fmt="#,##0")
    return r, r + 2


def build_personal_sheet(ws, author_disp, email, 정산서월, works, rsmap,
                         sales_path, etc_path, sales_cache=None, etc_cache=None):
    """개인(작가) 정산서. 원천징수 소득세 3% + 지방소득세 0.3%(소득세의 10%)."""
    y, mo = 정산서월.split("-")
    sheet = B._ymd_to_sheet(B._next_month(정산서월))
    웹툰, 기타 = [], []
    for w in works:
        rs = rsmap.get(w, 0.1)
        for d in B.연재상세(sales_path, w, rs, 당월_지급시트=sheet, cache=sales_cache):
            d["작품"] = w
            웹툰.append(d)
        for d in B.광고상세(etc_path, w, rs, 원작사정산월_cutoff=None, 당월only=정산서월, cache=etc_cache):
            d["작품"] = w
            기타.append(d)

    ws.cell(2, 2, f"{y}년 {mo}월 웹툰 원작료 정산서").font = Font(
        name="맑은 고딕", size=14, bold=True)

    # 섹션 먼저 그려 합계행 확보
    sum1, nxt = _detail_block(ws, 8, "1. 웹툰 원작료 정산 상세내역", 9, 웹툰, "작품명")
    sum2, _ = _detail_block(ws, nxt, "2. 기타 원작료 정산 상세내역", nxt + 1, 기타, "광고명")

    # 증빙표(행5 헤더, 행6 값)
    for c, t in {2: "정산문의", 5: "원작가명", 6: "작품명", 8: "증빙서류",
                 9: "총 정산금액", 10: "소득세", 11: "지방소득세", 12: "실지급액"}.items():
        _wcell(ws, 5, c, t, _BOLD, _CEN).fill = _HDRFILL
    _wcell(ws, 6, 2, email, _NORM, _LEFT)
    _wcell(ws, 6, 5, author_disp, _NORM, _CEN)
    _wcell(ws, 6, 6, works[0] if works else "", _NORM, _LEFT)
    _wcell(ws, 6, 8, "정산서", _NORM, _CEN)
    _wcell(ws, 6, 9, f"=K{sum1}+K{sum2}", _NORM, _RIGHT, fmt="#,##0")
    _wcell(ws, 6, 10, "=ROUNDDOWN(I6*0.03,-1)", _NORM, _RIGHT, fmt="#,##0")    # 소득세 3%
    _wcell(ws, 6, 11, "=ROUNDDOWN(J6*0.1,-1)", _NORM, _RIGHT, fmt="#,##0")     # 지방 0.3%
    _wcell(ws, 6, 12, "=I6-J6-K6", _BOLD, _RIGHT, fmt="#,##0")

    widths = {2: 22, 3: 8, 4: 9, 5: 16, 6: 18, 7: 11, 8: 11, 9: 14, 10: 11, 11: 11, 12: 13}
    for c, wd in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = wd
    return {"웹툰합": sum1, "기타합": sum2}


# ── 해외(Revenue Report): 도쿠마/신죠샤 — KRW 매출 → JPY, 누적 이력 ──
_COUNTRY_EN = {"북미": "North America", "미국": "North America", "스페인": "Spain",
               "이탈리아": "Italy", "대만": "Taiwan", "포르투갈": "Portugal",
               "홍콩": "Hong Kong", "일본": "Japan", "국내": ""}


def _platform_en(플랫폼, 국가):
    p = str(플랫폼 or "")
    if "카카오페이지" in p or "kakao" in p.lower():
        return "kakaopage"
    if "투믹스" in p or "toomics" in p.lower():
        ce = _COUNTRY_EN.get(str(국가 or ""), "")
        return f"toomics({ce})" if ce else "toomics"
    if "구루" in p:
        return "GRU COMPANY"
    if "코미코" in p:
        return "comico"
    if "만타" in p:
        return "MANTA"
    return p


def build_overseas_sheet(ws, vendor_name, 정산서월, works, rsmap, fx, 통화,
                         sales_path, sales_cache=None, etc_path=None, etc_cache=None,
                         제작비=0, period_months=None, title_en=None, 발송일=None,
                         인보이스_임계=0, 누적시작월=None):
    """해외 Revenue Report(도쿠마·신죠샤). 누적 이력(서비스월 ≤ 정산서월).
       인보이스_임계>0(도쿠마 10만엔): 누적 RS가 임계 초과 시에만 '발행 요청', 미만이면 누적 중.
       누적시작월: 직전 인보이스 후 리셋 기준월(이상부터 누적). None이면 전체이력."""
    title_en = title_en or {}
    rows = []
    for w in works:
        rs = rsmap.get(w, 0.1)
        for d in B.연재상세(sales_path, w, rs, 누적_cutoff=정산서월, cache=sales_cache):
            d["작품"] = w
            rows.append(d)
        if etc_path:
            for d in B.광고상세(etc_path, w, rs, 원작사정산월_cutoff=정산서월, cache=etc_cache):
                d["작품"] = w
                d["구분"] = "광고수익"
                rows.append(d)
    if 누적시작월:
        rows = [d for d in rows if d["서비스월"] and d["서비스월"] >= 누적시작월]
    rows.sort(key=lambda x: (x["작품"], x["서비스월"], str(x["플랫폼"])))
    warn = []
    py_rs = 0.0                              # 누적 RS(외화) 임계 판정용

    ws.cell(2, 2, "Revenue Report").font = Font(name="맑은 고딕", size=14, bold=True)
    ws.cell(2, 9, vendor_name).font = _BOLD
    # 1. Invoice Information
    _wcell(ws, 4, 2, "1. Invoice Information", _BOLD, _LEFT, border=False)
    for c, t in {2: "Month(Sending Report)", 5: "RS", 6: "Tax", 7: "RS-Tax", 8: "Currency"}.items():
        _wcell(ws, 6, c, t, _BOLD, _CEN).fill = _HDRFILL
    # 2. Accounts Detail
    _wcell(ws, 13, 2, "2. Accounts Detail", _BOLD, _LEFT, border=False)
    _wcell(ws, 13, 13, "Exchange Rate : Monthly average sales base rate (http://www.smbs.biz)",
           _NORM, _LEFT, border=False)
    제작비모드 = (제작비 or 0) > 0
    if 제작비모드:
        # 신죠샤: Title(EN) 없음 + 잔여제작비 열, 매월 RS만큼 차감
        DET = ["Month", "Title(KR)", "Platform", "Service Open", "Revenue\n(KRW)",
               "Rate", f"Revenue\n({통화})", "RS(%)", f"RS\n({통화})",
               f"Tax\n({통화}, 10%)", f"RS-Tax\n({통화})", f"잔여제작비(残余製作費)\n({통화})"]
        for c, t in enumerate(DET, 2):
            _wcell(ws, 15, c, t, _BOLD, _CEN).fill = _HDRFILL
        _wcell(ws, 16, 2, "제작화수(制作話数) 등 제작비", _NORM, _LEFT)
        _wcell(ws, 16, 13, 제작비, _NORM, _RIGHT, fmt="#,##0")
        first = 17
        r = first
        for d in rows:
            mon = d["서비스월"]; rate = B.fx_rate(fx, 통화, mon)
            plat_en = _platform_en(d["플랫폼"], d["국가"])
            if d.get("구분") == "광고수익":
                plat_en += "(advertising)"
            _wcell(ws, r, 2, mon, _NORM, _CEN)
            _wcell(ws, r, 3, d["작품"], _NORM, _LEFT)
            _wcell(ws, r, 4, plat_en, _NORM, _LEFT)
            _wcell(ws, r, 5, d.get("런칭") or "-", _NORM, _CEN)
            _wcell(ws, r, 6, d["정산기준순매출"], _NORM, _RIGHT, fmt="#,##0")
            if rate:
                _wcell(ws, r, 7, rate, _NORM, _CEN, fmt="0.0000")
                _wcell(ws, r, 8, f"=F{r}/G{r}", _NORM, _RIGHT, fmt="#,##0.00")
                py_rs += round(d["정산기준순매출"] / rate * d["RS"], 2)
            else:
                _wcell(ws, r, 7, "환율미입력", _NORM, _CEN)
                _wcell(ws, r, 8, 0, _NORM, _RIGHT); warn.append(mon)
            _wcell(ws, r, 9, d["RS"], _NORM, _CEN, fmt="0.0%")
            _wcell(ws, r, 10, f"=ROUND(H{r}*I{r},2)", _NORM, _RIGHT, fmt="#,##0.00")
            _wcell(ws, r, 11, f"=ROUND(J{r}*10%,0)", _NORM, _RIGHT, fmt="#,##0")
            _wcell(ws, r, 12, f"=J{r}-K{r}", _NORM, _RIGHT, fmt="#,##0.00")
            _wcell(ws, r, 13, f"=M{r-1}-J{r}", _NORM, _RIGHT, fmt="#,##0")
            r += 1
        last = r - 1
        _wcell(ws, r, 2, "Total", _BOLD, _CEN)
        _wcell(ws, r, 6, f"=SUBTOTAL(9,F{first}:F{last})", _BOLD, _RIGHT, fmt="#,##0")
        _wcell(ws, r, 10, f"=SUBTOTAL(9,J{first}:J{last})", _BOLD, _RIGHT, fmt="#,##0.00")
        _wcell(ws, r, 13, f"=M{last}", _BOLD, _RIGHT, fmt="#,##0")
        total_row = r
        잔여 = 제작비 - py_rs
        발행 = 잔여 <= 0
        _wcell(ws, 7, 2, 발송일 or _last_day(int(정산서월[:4]), int(정산서월[5:7])), _NORM, _CEN, fmt="yyyy.mm.dd")
        inv = 0 if 잔여 > 0 else round(-잔여, 0)
        _wcell(ws, 7, 5, inv, _NORM, _RIGHT, fmt="#,##0")
        _wcell(ws, 7, 6, 0, _NORM, _RIGHT, fmt="#,##0")
        _wcell(ws, 7, 7, inv, _NORM, _RIGHT, fmt="#,##0")
        _wcell(ws, 7, 8, 통화, _NORM, _CEN)
        _wcell(ws, 10, 2, "Add", _BOLD, _CEN)
        _wcell(ws, 10, 5, f"=J{total_row}", _BOLD, _RIGHT, fmt="#,##0.00")
        _wcell(ws, 10, 8, 통화, _BOLD, _CEN)
        return {"warn": sorted(set(warn)), "total_row": total_row, "rows": len(rows),
                "누적RS": py_rs, "발행": 발행, "임계": 인보이스_임계, "잔여제작비": round(잔여)}

    DET = ["Service Month", "Title(KR)", "Title(EN)", "Platform", "Service Open",
           "Revenue\n(KRW)", "Rate", f"Revenue\n({통화})", "RS(%)", f"RS\n({통화})",
           f"Tax\n({통화}, 10%)", f"RS-Tax\n({통화})"]
    for c, t in enumerate(DET, 2):
        _wcell(ws, 15, c, t, _BOLD, _CEN).fill = _HDRFILL
    first = 16
    r = first
    for d in rows:
        mon = d["서비스월"]
        rate = B.fx_rate(fx, 통화, mon)
        plat_en = _platform_en(d["플랫폼"], d["국가"])
        if d.get("구분") == "광고수익":
            plat_en += "(advertising)"
        _wcell(ws, r, 2, mon, _NORM, _CEN)
        _wcell(ws, r, 3, d["작품"], _NORM, _LEFT)
        _wcell(ws, r, 4, title_en.get(d["작품"], ""), _NORM, _LEFT)
        _wcell(ws, r, 5, plat_en, _NORM, _LEFT)
        _wcell(ws, r, 6, d.get("런칭") or "-", _NORM, _CEN)
        _wcell(ws, r, 7, d["정산기준순매출"], _NORM, _RIGHT, fmt="#,##0")
        if rate:
            _wcell(ws, r, 8, rate, _NORM, _CEN, fmt="0.0000")
            _wcell(ws, r, 9, f"=G{r}/H{r}", _NORM, _RIGHT, fmt="#,##0.00")
            py_rs += round(d["정산기준순매출"] / rate * d["RS"], 2)
        else:
            _wcell(ws, r, 8, "환율미입력", _NORM, _CEN)
            _wcell(ws, r, 9, 0, _NORM, _RIGHT)
            warn.append(mon)
        _wcell(ws, r, 10, d["RS"], _NORM, _CEN, fmt="0.0%")
        _wcell(ws, r, 11, f"=ROUND(I{r}*J{r},2)", _NORM, _RIGHT, fmt="#,##0.00")
        _wcell(ws, r, 12, f"=ROUND(K{r}*10%,0)", _NORM, _RIGHT, fmt="#,##0")
        _wcell(ws, r, 13, f"=K{r}-L{r}", _NORM, _RIGHT, fmt="#,##0.00")
        r += 1
    last = r - 1
    _wcell(ws, r, 2, "Total", _BOLD, _CEN)
    _wcell(ws, r, 7, f"=SUBTOTAL(9,G{first}:G{last})", _BOLD, _RIGHT, fmt="#,##0")
    _wcell(ws, r, 9, f"=SUBTOTAL(9,I{first}:I{last})", _BOLD, _RIGHT, fmt="#,##0.00")
    _wcell(ws, r, 11, f"=SUBTOTAL(9,K{first}:K{last})", _BOLD, _RIGHT, fmt="#,##0.00")
    _wcell(ws, r, 12, f"=SUBTOTAL(9,L{first}:L{last})", _BOLD, _RIGHT, fmt="#,##0")
    _wcell(ws, r, 13, f"=SUBTOTAL(9,M{first}:M{last})", _BOLD, _RIGHT, fmt="#,##0.00")
    total_row = r

    # Invoice 값(합계 참조) + 10만엔 임계 판정
    발행 = (인보이스_임계 <= 0) or (py_rs > 인보이스_임계)
    _wcell(ws, 7, 2, 발송일 or _last_day(int(정산서월[:4]), int(정산서월[5:7])), _NORM, _CEN, fmt="yyyy.mm.dd")
    _wcell(ws, 7, 5, f"=K{total_row}", _NORM, _RIGHT, fmt="#,##0.00")
    _wcell(ws, 7, 6, f"=L{total_row}", _NORM, _RIGHT, fmt="#,##0")
    _wcell(ws, 7, 7, f"=M{total_row}", _NORM, _RIGHT, fmt="#,##0.00")
    _wcell(ws, 7, 8, 통화, _NORM, _CEN)
    _wcell(ws, 10, 2, "Add", _BOLD, _CEN)
    _wcell(ws, 10, 5, f"=K{total_row}", _BOLD, _RIGHT, fmt="#,##0.00")
    _wcell(ws, 10, 6, f"=L{total_row}", _BOLD, _RIGHT, fmt="#,##0")
    _wcell(ws, 10, 7, f"=M{total_row}", _BOLD, _RIGHT, fmt="#,##0.00")
    _wcell(ws, 10, 8, 통화, _BOLD, _CEN)
    if 인보이스_임계 > 0:
        if 발행:
            msg = (f"누적 RS {py_rs:,.0f} {통화} > {인보이스_임계:,} {통화} → 인보이스 발행 요청"
                   f" (Invoice requested)")
        else:
            msg = (f"누적 RS {py_rs:,.0f} {통화} ≤ {인보이스_임계:,} {통화} → 누적 중, 인보이스 미발행"
                   f" (Below threshold, carried forward)")
        c = ws.cell(11, 2, msg)
        c.font = Font(name="맑은 고딕", size=10, bold=True,
                      color="C00000" if 발행 else "808080")

    widths = {2: 13, 3: 20, 4: 26, 5: 18, 6: 11, 7: 12, 8: 9, 9: 12, 10: 7, 11: 11, 12: 12, 13: 12}
    for c, wd in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = wd
    return {"warn": warn, "total_row": total_row, "rows": len(rows),
            "누적RS": round(py_rs, 2), "발행": 발행, "임계": 인보이스_임계}


def build_quarterly_sheet(ws, vendor_name, 정산서월, works, rsmap,
                          sales_path, sales_cache=None, period_months=None, 품목_label=None,
                          작성일=None, 마감일=None, 지급일=None,
                          발송이메일="terapin_toonbill@terapinstudios.co.kr"):
    """리디 분기 정산서. 복합 외화 표시(총매출 외화/원화·정산금 외화·tax·환율·정산기준순매출 외화/원화).
       RAW는 원화만 제공 → 원화 컬럼·정산금액 자동, 외화 컬럼은 추후 기재(빈칸).
       정산금액 = ROUND(정산기준순매출 원화 × 원작사 RS). 작품별 그룹핑(원본 구조)."""
    label = 품목_label or 정산서월
    # 제목: "2026.2Q" → "2026년 2Q 리디 웹툰 원작료 정산서"
    _t = str(label).replace(".", "년 ", 1) if "Q" in str(label) else label
    title = f"{_t} {vendor_name} 웹툰 원작료 정산서"
    ws.cell(2, 2, title).font = Font(name="맑은 고딕", size=13, bold=True)
    # ── 증빙표 (리디 양식: 공급가액 G / 부가세 I / 실지급 K / 이메일 M) ──
    for c, t in {2: "증빙서류", 4: "작성일자", 5: "품목명", 7: "공급가액",
                 9: "부가세", 11: "실지급액", 13: "(세금)계산서 발송 이메일 주소",
                 14: "마감일", 15: "지급일"}.items():
        _wcell(ws, 5, c, t, _BOLD, _CEN).fill = _HDRFILL
    works_disp = ", ".join(works)
    _wcell(ws, 6, 2, "세금계산서", _NORM, _CEN)
    if 작성일:
        _wcell(ws, 6, 4, 작성일, _NORM, _CEN, fmt="yyyy-mm-dd")
    _wcell(ws, 6, 5, f"원작료_{label} 정산 <{works_disp}>", _NORM, _LEFT)
    # 공급가액 = 정산금액 합(상세 Q열 합계) — 상세 합계행 참조는 아래에서 채움
    _wcell(ws, 6, 9, "=ROUND(G6*10%,0)", _NORM, _RIGHT, fmt="#,##0")
    _wcell(ws, 6, 11, "=G6+I6", _NORM, _RIGHT, fmt="#,##0")
    _wcell(ws, 6, 13, 발송이메일, _NORM, _LEFT)
    if 마감일:
        _wcell(ws, 6, 14, 마감일, _NORM, _CEN, fmt="yyyy-mm-dd")
    if 지급일:
        _wcell(ws, 6, 15, 지급일, _NORM, _CEN, fmt="yyyy-mm-dd")
    _wcell(ws, 7, 2, "합계", _BOLD, _CEN)
    _wcell(ws, 7, 7, "=SUM(G6)", _BOLD, _RIGHT, fmt="#,##0")
    _wcell(ws, 7, 9, "=SUM(I6)", _BOLD, _RIGHT, fmt="#,##0")
    _wcell(ws, 7, 11, "=SUM(K6)", _BOLD, _RIGHT, fmt="#,##0")
    _wcell(ws, 7, 13, "(세금)계산서는 1장으로 발행해 주시되 품목 나눠 기재 부탁드립니다.", _NORM, _LEFT)

    # ── 상세 (17컬럼) : 작품별 섹션 ──
    DET = {2: "순번", 3: "국가", 4: "구분", 5: "플랫폼명", 6: "작품명", 7: "런칭일",
           8: "서비스월", 9: "총매출\n(JPY/USD)", 10: "총매출\n(원화)", 11: "정산금\n(JPY/USD)",
           12: "tax\n(JPY, 10%)", 13: "환율", 14: "정산기준순매출\n(JPY/USD)",
           15: "정산기준순매출\n(원화)", 16: "원작사 RS율", 17: "정산금액"}
    # 작품별 데이터 수집(분기 내 서비스월, 월·국가·플랫폼 정렬)
    rows_by_work = {}
    for w in works:
        rs = rsmap.get(w, 0.1)
        rws = [dict(d, 작품=w) for d in
               B.연재상세(sales_path, w, rs, 서비스월_set=period_months, cache=sales_cache)]
        rws.sort(key=lambda x: (str(x["서비스월"]), x["국가"], str(x["플랫폼"])))
        if rws:
            rows_by_work[w] = rws

    r = 9
    sub_cells = []                      # 작품별 합계 셀(증빙표 공급가 합산용)
    sec_no = 0
    for w in works:
        rws = rows_by_work.get(w)
        if not rws:
            continue
        sec_no += 1
        _wcell(ws, r, 2, f"{sec_no}. 웹툰 <{w}> 원작료 정산 상세내역", _BOLD, _LEFT, border=False)
        r += 1
        for c, t in DET.items():        # 섹션마다 헤더
            _wcell(ws, r, c, t, _BOLD, _CEN).fill = _HDRFILL
        r += 1
        first = r
        seq = 0
        for d in rws:
            seq += 1
            _wcell(ws, r, 2, seq, _NORM, _CEN)
            _wcell(ws, r, 3, d["국가"], _NORM, _CEN)
            _wcell(ws, r, 4, d["구분"], _NORM, _CEN)
            _wcell(ws, r, 5, d["플랫폼"], _NORM, _LEFT)
            _wcell(ws, r, 6, d["작품"], _NORM, _LEFT)
            _wcell(ws, r, 7, d.get("런칭") or "-", _NORM, _CEN)
            _wcell(ws, r, 8, d["서비스월"], _NORM, _CEN)
            # 외화 컬럼(9,11,12,13,14)은 추후 기재. 원화(10)·정산기준순매출 원화(15)는 RAW.
            _wcell(ws, r, 10, d.get("총매출") or 0, _NORM, _RIGHT, fmt="#,##0")
            _wcell(ws, r, 15, d["정산기준순매출"], _NORM, _RIGHT, fmt="#,##0")
            _wcell(ws, r, 16, d["RS"], _NORM, _CEN, fmt="0.0%")
            _wcell(ws, r, 17, f"=ROUND(O{r}*P{r},0)", _NORM, _RIGHT, fmt="#,##0")
            r += 1
        last = r - 1
        _wcell(ws, r, 2, "합계", _BOLD, _CEN)
        _wcell(ws, r, 17, f"=SUM(Q{first}:Q{last})", _BOLD, _RIGHT, fmt="#,##0")
        sub_cells.append(f"Q{r}")
        r += 2                          # 작품 섹션 사이 공백

    # 증빙표 공급가액 = 모든 작품 정산금액 합계
    ws.cell(6, 7, ("=" + "+".join(sub_cells)) if sub_cells else 0)
    ws.cell(6, 7).number_format = "#,##0"
    widths = {2: 5, 3: 8, 4: 8, 5: 14, 6: 14, 7: 11, 8: 11, 9: 11, 10: 12,
              11: 11, 12: 10, 13: 9, 14: 13, 15: 13, 16: 10, 17: 12}
    for c, wd in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = wd
    return works


def make_personal(template_path, out_path, vendor, email, 정산서월, works, rsmap,
                  sales_path, etc_path, sales_cache=None, etc_cache=None):
    """테라핀_개인 템플릿을 채워 개인(작가) 정산서 생성(데이터 수집 포함)."""
    import template_fill as TF
    y, mo = 정산서월.split("-")
    sheet = B._ymd_to_sheet(B._next_month(정산서월))
    웹툰, 기타 = [], []
    for w in works:
        rs = rsmap.get(w, 0.1)
        for d in B.연재상세(sales_path, w, rs, 당월_지급시트=sheet, cache=sales_cache):
            d["작품"] = w
            웹툰.append(d)
        for d in B.광고상세(etc_path, w, rs, 원작사정산월_cutoff=None,
                          당월only=정산서월, cache=etc_cache):
            d["작품"] = w
            기타.append(d)
    ctx = {"year": y, "month": mo, "author_disp": vendor, "email": email,
           "작품명": works[0] if works else "", "웹툰": 웹툰, "기타": 기타}
    TF.fill_personal(template_path, out_path, ctx)
    return {"웹툰수": len(웹툰), "기타수": len(기타)}
