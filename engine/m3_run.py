# -*- coding: utf-8 -*-
"""
M3 오케스트레이터 — 회사·정산서월 + RAW(매출/기타수익) + 마스터 → 그룹별 정산서 생성.

그룹: 네이버 / 리디·분기 / 그외 / 개인 / 해외 / 수성
출력: output/YYYY-MM(정산서월)/회사/원작료정산서_<그룹>_YYYYMM.xlsx (업체=시트)

현 단계: 사업자(세금계산서) 양식 — 네이버·그외 월간 사업자. MG·이월(봄봄) 반영.
개인/해외(Revenue Report)/리디분기/반기/수성은 전용 양식으로 후속 확장.
"""
import os
from collections import defaultdict
from openpyxl import Workbook, load_workbook

import m3_builder as B
import m3_sheet as S


def _template_path():
    """정산서 양식 템플릿 경로. paths.template_path() 우선, 없으면 루트의 기초 파일."""
    try:
        import sys
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if root not in sys.path:
            sys.path.insert(0, root)
        import paths as _p
        return _p.template_path()
    except Exception:
        return os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            "정산서_양식_기초.xlsx")


def _fill_business(out_path, ctx):
    import sys
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if root not in sys.path:
        sys.path.insert(0, root)
    import template_fill as TF
    return TF.fill_business(_template_path(), out_path, ctx)


def _load_masters(master_path):
    wb = load_workbook(master_path, data_only=True)
    al, vm, mg = wb["작품Alias마스터"], wb["업체마스터"], wb["선인세MG마스터"]
    vinfo, vworks, rsmap = {}, defaultdict(list), {}
    title_en = {}
    wsplit = {}                                 # 작품 → 정산서분리 태그
    ah = {al.cell(3, c).value: c for c in range(1, al.max_column + 1) if al.cell(3, c).value}
    split_col = ah.get("정산서분리")
    vh = {vm.cell(3, c).value: c for c in range(1, vm.max_column + 1) if vm.cell(3, c).value}

    def _vc(r, header, default=None):
        c = vh.get(header)
        v = vm.cell(r, c).value if c else None
        return v if v is not None else default

    for r in range(4, vm.max_row + 1):
        vid = vm.cell(r, 1).value
        if str(vid).startswith("V"):
            vinfo[str(vid)] = dict(
                명=_vc(r, "업체명"),
                주체=_vc(r, "정산주체(회사)") or _vc(r, "정산주체") or _vc(r, "주체"),
                유형=_vc(r, "유형"), 소속=_vc(r, "소속"),
                정산유형=_vc(r, "정산유형"),
                증빙=_vc(r, "증빙구분", "세금계산서"),
                발송월=_vc(r, "발송월", "매월"),
                발송오프셋=_vc(r, "발송오프셋", 0),
                발송일=_vc(r, "발송일", "말일"))
    for r in range(4, al.max_row + 1):
        nm, vid = al.cell(r, 1).value, str(al.cell(r, 5).value)
        if nm and vid.startswith("V") and "예시" not in str(nm):
            vworks[vid].append(str(nm))
            try:
                rsmap[str(nm)] = float(al.cell(r, 6).value)
            except (TypeError, ValueError):
                rsmap[str(nm)] = 0.1
            en = al.cell(r, 13).value           # 영문제목(선택)
            if en:
                title_en[str(nm)] = str(en)
            if split_col:                        # 정산서분리 태그(선택)
                tag = al.cell(r, split_col).value
                if tag and str(tag).strip():
                    wsplit[str(nm)] = str(tag).strip()
    mgset = set()
    for r in range(4, mg.max_row + 1):
        if mg.cell(r, 2).value and mg.cell(r, 3).value and "예시" not in str(mg.cell(r, 2).value):
            mgset.add(str(mg.cell(r, 2).value))
    return vinfo, vworks, rsmap, mgset, title_en, wsplit


def _group_of(vi):
    if vi.get("주체") == "수성웹툰":
        return "수성"
    if vi.get("소속") == "해외":
        return "해외"
    nm = str(vi.get("명") or "")
    if "네이버" in nm:
        return "네이버"
    if nm.startswith("리디") or vi.get("정산유형") == "분기":
        return "리디분기"
    if vi.get("유형") == "개인":
        return "개인"
    return "그외"


_PRODUCTION_COST = {"V034": 7500000}  # 신죠샤 밑바닥 마술사 제작비(JPY) — 운영 시 마스터 이관


def run_settlement(master_path, sales_path, etc_path, company, 정산서월,
                   out_root="output", groups=("네이버", "그외"), 이월_prior=None):
    """company의 정산서월 정산서를 그룹별로 생성. 검수 리포트(report) 함께 반환.
       이월_prior: {업체ID: 직전정산서_경로} — 이월 합산 시 직전 정산기준순매출 자동 산출."""
    vinfo, vworks, rsmap, mgset, title_en, wsplit = _load_masters(master_path)
    print("매출·기타수익 캐시 로딩...", flush=True)
    sales_cache = B.load_sales(sales_path)
    etc_cache = B.load_etc(etc_path)
    print(f"  매출 작품 {len(sales_cache)} · 기타 작품 {len(etc_cache)}", flush=True)
    rs_override = _tier_rs_overrides(master_path, 정산서월, sales_cache, rsmap)
    carryover = _load_carryover(master_path, 정산서월)
    이월_prior = 이월_prior or {}
    for vid, ents in carryover.items():          # 합산 건 이월순매출 자동 산출(직전 정산서)
        ppath = 이월_prior.get(vid)
        for w, e in ents.items():
            if e.get("증빙상태") == "수취" and e.get("합산월") == 정산서월 and ppath:
                net = _prior_settlement_net(ppath, w, _prev_month(정산서월))
                if net is not None:
                    e["이월순매출"] = net
    if rs_override:
        print(f"  누적구간 tier RS: {rs_override}", flush=True)
    fx = B.load_fx(master_path)

    bygroup = defaultdict(list)
    for vid, works in vworks.items():
        vi = vinfo.get(vid, {})
        주체 = vi.get("주체")
        if company == "수성웹툰" and 주체 != "수성웹툰":
            continue
        if company == "테라핀" and 주체 == "수성웹툰":
            continue
        bygroup[_group_of(vi)].append((vid, vi, works))

    out_dir = os.path.join(out_root, 정산서월, company)
    os.makedirs(out_dir, exist_ok=True)
    produced = []
    report = dict(company=company, 정산서월=정산서월, files=[], 업체수=0,
                  처리작품수=0, MG작품수=0, 해외정산수=0, 이월대상수=0,
                  오류건수=0, 경고=[], 미구현그룹=[])
    drawn_works = set()
    y, mo = 정산서월.split("-")
    for g in groups:
        vendors = bygroup.get(g, [])
        if not vendors:
            continue
        gdir = os.path.join(out_dir, g)
        os.makedirs(gdir, exist_ok=True)
        made = 0
        for vid, vi, works in sorted(vendors):
            유형 = vi.get("유형")
            정산유형 = vi.get("정산유형")
            send, period, label = _send_and_period(정산서월, 정산유형,
                                                    vi.get("발송월"), vi.get("발송오프셋"))
            if not send:
                report["경고"].append(
                    f"{vi.get('명')}({vid}) {정산유형} - 이번 달 발송월 아님, 생략")
                continue
            wb = Workbook()
            ws = wb.active
            ws.title = _safe_sheet(label or 정산서월)
            tpl_path = None
            if g == "해외":
                통화 = "JPY"
                임계 = 100000 if 정산유형 == "분기" else 0   # 도쿠마: 누적 10만엔 초과 시 발행
                제작비 = _PRODUCTION_COST.get(vid, 0)        # 신죠샤: 제작비(JPY) 회수 모델
                o = S.build_overseas_sheet(ws, vi["명"], 정산서월, works, rsmap, fx, 통화,
                                           sales_path, sales_cache=sales_cache,
                                           etc_path=etc_path, etc_cache=etc_cache,
                                           title_en=title_en, 인보이스_임계=임계, 제작비=제작비)
                if o.get("warn"):
                    report["경고"].append(f"{vi['명']} 환율 미입력월: {sorted(set(o['warn']))}")
                if 제작비 and not o.get("발행"):
                    report["경고"].append(
                        f"{vi['명']} 잔여제작비 {o.get('잔여제작비'):,} JPY > 0 → Invoice 0(제작비 회수중)")
                elif 임계 and not o.get("발행"):
                    report["경고"].append(
                        f"{vi['명']} 누적 RS {o['누적RS']:,.0f} JPY ≤ {임계:,} → 인보이스 미발행(누적)")
                sec = works if o.get("rows") else None
            elif 유형 == "개인":
                email = _recipient_of(master_path, vid) or _email_of(master_path, vid)
                p = os.path.join(gdir, f"원작료정산서_{_safe_name(vi['명'])}_{y}{mo}.xlsx")
                S.make_personal(_template_path(), p, vi["명"], email, 정산서월,
                                works, rsmap, sales_path, etc_path,
                                sales_cache=sales_cache, etc_cache=etc_cache)
                made += 1
                report["업체수"] += 1
                report["files"].append(dict(group=g, vid=vid, 업체=vi["명"], path=p))
                drawn_works.update(works)
                continue
            elif g == "리디분기":
                발행, 마감, 지급 = _doc_dates("리디", 정산서월)
                sec = S.build_quarterly_sheet(ws, vi["명"], 정산서월, works, rsmap,
                                              sales_path, sales_cache=sales_cache,
                                              period_months=period, 품목_label=label,
                                              작성일=발행, 마감일=마감, 지급일=지급)
            elif 유형 == "사업자":
                cat = _vendor_cat(vi, g)
                발행, 마감, 지급 = _doc_dates(cat, 정산서월)
                직접 = (cat == "카카오")
                if cat == "카카오":
                    # 카카오: 단일 업체이나 플랫폼별(카카오웹툰·카카오페이지·타파스_미국·픽코마_일본)로 분할
                    ctx = S.collect_vendor_ctx(
                        vi["명"], 정산서월, works, rsmap, mgset, sales_path, etc_path,
                        master_path, sales_cache=sales_cache, etc_cache=etc_cache,
                        rs_override=rs_override, period_months=period, 품목_label=label,
                        이월_map=carryover.get(vid, {}), 직접정산=True,
                        작성일=발행, 마감일=마감, 지급일=지급,
                        증빙구분=(vi.get("증빙") or "세금계산서"))
                    if not ctx["works"]:
                        continue
                    made_here = 0
                    for glabel, sub in _split_ctx_by_kakao(ctx):
                        p = os.path.join(gdir, f"원작료정산서_카카오_{_safe_name(glabel)}_{y}{mo}.xlsx")
                        _fill_business(p, sub)
                        made_here += 1
                        report["업체수"] += 1
                        report["files"].append(dict(group=g, vid=vid,
                                                    업체=f"카카오 / {glabel}", path=p))
                        for wk in sub["works"]:
                            drawn_works.add(wk["작품"])
                    made += made_here
                    continue
                # 그 외 사업자: 정산서분리 태그별로 그룹 분할(태그 없으면 본 정산서 1건)
                buckets = defaultdict(list)
                for w in works:
                    buckets[wsplit.get(w, "")].append(w)
                made_here = 0
                for tag, gworks in sorted(buckets.items()):
                    ctx = S.collect_vendor_ctx(
                        vi["명"], 정산서월, gworks, rsmap, mgset, sales_path, etc_path,
                        master_path, sales_cache=sales_cache, etc_cache=etc_cache,
                        rs_override=rs_override, period_months=period, 품목_label=label,
                        이월_map=carryover.get(vid, {}), 직접정산=직접,
                        작성일=발행, 마감일=마감, 지급일=지급,
                        증빙구분=(vi.get("증빙") or "세금계산서"))
                    if not ctx["works"]:
                        continue
                    suffix = f"_{_safe_name(tag)}" if tag else ""
                    p = os.path.join(gdir, f"원작료정산서_{_safe_name(vi['명'])}{suffix}_{y}{mo}.xlsx")
                    _fill_business(p, ctx)
                    made_here += 1
                    report["업체수"] += 1
                    report["files"].append(dict(group=g, vid=vid,
                                                업체=vi["명"] + (f" / {tag}" if tag else ""), path=p))
                    for wk in ctx["works"]:
                        drawn_works.add(wk["작품"])
                made += made_here
                continue                          # 저장·report는 위에서 처리 완료
            else:
                report["경고"].append(f"{vi.get('명')}({vid}) 유형 미지원 - 건너뜀")
                continue
            if not sec:
                continue
            if tpl_path:
                path = tpl_path                       # template_fill이 이미 저장
            else:
                path = os.path.join(gdir, f"원작료정산서_{_safe_name(vi['명'])}_{y}{mo}.xlsx")
                wb.save(path)
            made += 1
            report["업체수"] += 1
            report["files"].append(dict(group=g, vid=vid, 업체=vi["명"], path=path))
            for w in (sec if isinstance(sec, dict) else sec):
                if isinstance(w, str):
                    drawn_works.add(w)
            if 유형 == "개인":
                drawn_works.update(works)
        if made:
            produced.append((g, made))
            print(f"  [{g}] {made}개 업체 파일 생성", flush=True)

    # 미구현 그룹(전용 양식 대기) 기록
    for g in ("개인", "해외", "리디분기", "수성"):
        if g not in groups and bygroup.get(g):
            report["미구현그룹"].append(f"{g}({len(bygroup[g])}업체)")

    report["처리작품수"] = len(drawn_works)
    report["MG작품수"] = len(drawn_works & mgset)
    report["해외정산수"] = sum(len(w) for vid, vi, w in bygroup.get("해외", []))
    report["분기반기수"] = sum(1 for g in groups for (vid, vi, w) in bygroup.get(g, [])
                            if str(vi.get("정산유형")) in ("분기", "반기"))
    report["이월대상수"] = _carryover_count(master_path, 정산서월)
    return produced, report


def _carryover_count(master_path, 정산서월):
    """이월마스터에서 이번 정산서월에 관련된 이월(보류/합산) 건수."""
    try:
        ws = load_workbook(master_path, data_only=True)["이월마스터"]
    except KeyError:
        return 0
    cnt = 0
    for r in range(4, ws.max_row + 1):
        vid = ws.cell(r, 1).value
        if not str(vid).startswith("V") or "예시" in str(ws.cell(r, 2).value or ""):
            continue
        발생 = str(ws.cell(r, 3).value or "")
        합산 = str(ws.cell(r, 5).value or "")
        증빙 = str(ws.cell(r, 4).value or "")
        # 미수취(보류 진행) 또는 이번 합산월에 해소되는 건
        if 합산 == 정산서월 or (증빙 == "미수취" and 발생 <= 정산서월):
            cnt += 1
    return cnt


def _tier_rs_overrides(master_path, 정산서월, sales_cache, rsmap):
    """누적구간(만년 등) 작품의 당월 효과 연재RS 계산 → {작품: RS}.
       당월 정산기준에 ledger tier 적용 → won/정산기준."""
    import sys as _sys
    _sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    import ledger as LG
    wb = load_workbook(master_path, data_only=True)
    ws = wb["예외규칙마스터"]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    rules = {}
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, hdr.get("규칙유형", 0)).value) == "누적순매출구간":
            nm = ws.cell(r, hdr["표준작품명"]).value
            rules[str(nm)] = {k: ws.cell(r, c).value for k, c in hdr.items()}
    if not rules:
        return {}
    sheet = B._ymd_to_sheet(B._next_month(정산서월))
    prev_cum = None
    out = {}
    for nm, rl in rules.items():
        base = sum(d["정산기준"] for d in sales_cache.get(B._norm(nm), [])
                   if d["지급시트"] == sheet and "독점" in str(d["구분"] or ""))
        if base <= 0:
            continue
        tiers = LG.parse_tiers(rl)
        # 판정기준: 기본 '월별'(당월 정산대상 순매출). '누적'이면 ledger 누적전 사용.
        판정 = str(rl.get("판정기준") or "월별").strip()
        if 판정 == "누적":
            if prev_cum is None:
                prev_cum = LG.prev_cumulative(master_path, 정산서월)
            before = prev_cum.get(nm, 0.0)
        else:
            before = 0.0                    # 월별: 누적 무시, 당월 base로만 구간 판정
        won, _, _ = LG.tiered_원작료(base, base, before, rsmap.get(nm, 0.1), tiers,
                                   str(rl.get("RS적용base") or "정산기준매출"))
        out[nm] = round(won / base, 6)
    return out


def _safe_sheet(name):
    for ch in r'[]:*?/\\':
        name = name.replace(ch, "")
    return name[:31]


def _safe_name(name):
    """파일명용 업체명 정리."""
    for ch in r'[]:*?/\\<>|"':
        name = str(name).replace(ch, "")
    return name.strip()


def _email_of(master_path, vid):
    """이메일마스터에서 vid의 발신 주소(개인 정산문의용)."""
    try:
        ws = load_workbook(master_path, data_only=True)["이메일마스터"]
    except KeyError:
        return ""
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == str(vid):
            return ws.cell(r, 2).value or ""
    return ""


def _recipient_of(master_path, vid):
    """이메일마스터에서 vid의 수신 주소(C열) — 개인 정산서 원작가명 칸의 작가 메일."""
    try:
        ws = load_workbook(master_path, data_only=True)["이메일마스터"]
    except KeyError:
        return ""
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == str(vid):
            v = ws.cell(r, 3).value or ""
            return str(v).split(";")[0].strip()       # 첫 수신 주소
    return ""


def _parse_발송월(발송월):
    s = str(발송월 or "매월").strip()
    if s in ("매월", "", "None"):
        return None                      # 매월(월간)
    return {int(x) for x in s.replace(" ", "").split(",") if x.isdigit()}


def _send_and_period(정산서월, 정산유형, 발송월, 발송오프셋):
    """발송 여부·기간(서비스월 집합)·라벨 결정.
       발송월: 발송하는 월 집합(매월=None). 발송오프셋: 기간 종료월 = 정산서월-오프셋(개월).
       리디=1,4,7,10/오프셋1(직전분기), 도쿠마=3,6,9,12/오프셋0, 반기=6,12, 월간=매월."""
    y, m = map(int, 정산서월.split("-"))
    월set = _parse_발송월(발송월)
    if 월set is not None and m not in 월set:
        return False, None, None         # 이번 달은 발송 안 함
    off = int(발송오프셋 or 0)
    ey, em = y, m - off                  # 기간 종료월
    while em < 1:
        em += 12
        ey -= 1
    if 정산유형 == "분기":
        months = set()
        cy, cm = ey, em
        for _ in range(3):
            months.add(f"{cy}-{cm:02d}")
            cm -= 1
            if cm < 1:
                cm = 12
                cy -= 1
        return True, months, f"{ey}.{(em - 1) // 3 + 1}Q"
    if 정산유형 == "반기":
        start = 1 if em <= 6 else 7
        months = {f"{ey}-{mm:02d}" for mm in range(start, em + 1)}
        return True, months, f"{ey} {'상반기' if em <= 6 else '하반기'}"
    return True, None, None              # 월간(당월)


def _load_carryover(master_path, 정산서월):
    """이월마스터 → {업체ID: {작품: {증빙상태, 발생월, 합산월}}}.
       미수취(발생월≤정산서월, 미해소) 또는 이번 합산월(수취) 건만."""
    out = {}
    try:
        ws = load_workbook(master_path, data_only=True)["이월마스터"]
    except KeyError:
        return out
    for r in range(4, ws.max_row + 1):
        vid = str(ws.cell(r, 1).value or "")
        nm = str(ws.cell(r, 2).value or "")
        if not vid.startswith("V") or "예시" in nm:
            continue
        발생 = str(ws.cell(r, 3).value or "")
        증빙 = str(ws.cell(r, 4).value or "").strip()
        합산 = str(ws.cell(r, 5).value or "")
        active = (증빙 == "수취" and 합산 == 정산서월) or \
                 (증빙 == "미수취" and 발생 and 발생 <= 정산서월)
        if active:
            out.setdefault(vid, {})[nm] = {"증빙상태": 증빙, "발생월": 발생, "합산월": 합산}
    return out


def _prior_settlement_net(path, work, prior_month=None):
    """직전 정산서 파일에서 작품의 정산기준순매출(J/10열) 합 → 이월 정산금 기준값.
       prior_month 주면 해당 월 시트(예 '2026-04월'/'04월') 우선, 없으면 active."""
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return None
    ws = None
    if prior_month:
        mm = prior_month[5:7]
        for nm in wb.sheetnames:
            s = str(nm)
            if s in (f"{prior_month}월", f"{prior_month[:4]}-{mm}월", f"{mm}월") or prior_month in s:
                ws = wb[nm]
                break
    if ws is None:
        ws = wb.active
    # 헤더행(순번/국가...) 다음부터 작품 상세행의 J(10)열 합
    tot = 0.0
    for r in range(1, ws.max_row + 1):
        f = ws.cell(r, 6).value           # 작품명(F)
        j = ws.cell(r, 10).value          # 정산기준순매출(J)
        if f and str(f).strip() == str(work).strip() and isinstance(j, (int, float)):
            tot += j
    return round(tot)


def _prev_month(ym):
    y, m = int(ym[:4]), int(ym[5:7])
    return f"{y - 1}-12" if m == 1 else f"{y}-{m - 1:02d}"


def _prev_bizday(d):
    """주말이면 직전 영업일(금)로. (공휴일 리스트는 추후 마스터로 확장)"""
    import datetime as _dt
    while d.weekday() >= 5:               # 5=토,6=일
        d -= _dt.timedelta(days=1)
    return d


def _doc_dates(cat, 정산서월):
    """업체 구분별 (계산서 발행일, 마감일, 지급일). None은 '-'(해당없음).
       N=정산서월. 매월4일 발송군(수성·그외·카카오·개인)은 처리월=N+1,
       네이버·리디는 발행=N 말일·마감/지급=N+1."""
    import datetime as _dt, calendar
    y, m = int(정산서월[:4]), int(정산서월[5:7])
    ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)        # N+1
    def dN(day):  return _dt.datetime(y, m, day)
    def dN1(day): return _dt.datetime(ny, nm, day)
    lastN = _dt.datetime(y, m, calendar.monthrange(y, m)[1])
    if cat in ("수성", "그외"):
        return dN1(5), dN1(10), _prev_bizday(dN1(15))
    if cat == "네이버":
        return lastN, dN1(8), _prev_bizday(dN1(15))
    if cat == "리디":
        return lastN, dN1(8), dN1(10)
    if cat == "개인":
        return None, None, _prev_bizday(dN1(15))
    return None, None, None                                # 카카오·해외(도쿠마·신죠샤)


def _vendor_cat(vi, group):
    명 = str(vi.get("명", ""))
    if str(vi.get("회사") or vi.get("정산주체") or "") .startswith("수성"):
        return "수성"
    if 명.startswith("카카오"):
        return "카카오"
    if group == "네이버":
        return "네이버"
    if group in ("리디분기", "리디"):
        return "리디"
    if group == "개인":
        return "개인"
    if group == "해외":
        return "해외"
    return "그외"


def _kakao_group(국가, 플랫폼):
    """카카오 정산서를 플랫폼별로 분할하는 그룹 라벨."""
    p = str(플랫폼 or "")
    if "픽코마" in p or "piccoma" in p.lower():
        return "픽코마_일본"
    if "타파스" in p or "tapas" in p.lower():
        return "타파스_미국"
    if "카카오웹툰" in p:
        return "카카오웹툰"
    if "카카오페이지" in p:
        return "카카오페이지"
    return p or "카카오"


def _split_ctx_by_kakao(ctx):
    """카카오 ctx를 플랫폼 그룹별 sub-ctx 로 쪼갠다. [(라벨, sub_ctx), ...]"""
    buckets = {}
    for w in ctx["works"]:
        per = {}
        for rw in w.get("rows", []):
            g = _kakao_group(rw.get("국가"), rw.get("플랫폼"))
            per.setdefault(g, []).append(rw)
        for g, rows in per.items():
            wc = dict(w)
            wc["rows"] = rows
            buckets.setdefault(g, []).append(wc)
    out = []
    for g in buckets:
        sub = dict(ctx)
        sub["works"] = buckets[g]
        sub["vendor"] = g                 # 제목·품목에 플랫폼 라벨 사용
        out.append((g, sub))
    return out
