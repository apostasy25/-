# -*- coding: utf-8 -*-
"""
M0 입력 로더 — 전체 정산 엔진(방향 B)의 첫 단계.

역할: inbox/매출리스트, inbox/기타수익 의 엑셀을 읽어
      뒤 단계(M1 매칭·M2 금액)가 바로 쓸 수 있는 '표준 레코드'로 변환한다.

설계 원칙
  · 컬럼은 위치가 아니라 '키워드'로 인식한다.
    (파일마다 라벨/헤더행/공백이 달라도 견디게 — 예: 정산기준매출(A)/(B),
     '지급월'/'지급 월', 광고시트 헤더 6행 vs IP시트 5행)
  · 합계/소계/빈 행은 자동 제외.
  · 회사(테라핀/수성)·정산월은 파일명에서 추출.
  · 읽기만 한다(원본 무수정). 계산·매칭은 M1/M2 담당.
"""
import os
import re
import glob
from openpyxl import load_workbook


# ── 헤더 키워드 매핑 ─────────────────────────────────────────
def _t(h):
    return str(h).replace(" ", "").replace("\n", "") if h is not None else ""


def _canon_sales(h):
    """매출리스트 헤더 1칸 → 표준 필드명."""
    t = _t(h)
    # 원작료(C) 열 헤더는 "(B×원작사 RS율)" 설명을 품어 RS/원작사 키워드가 섞이므로 먼저 확정
    if "원작료" in t:
        return "원작료"
    if "RS율" in t:                                  # RS율이 둘(테라핀/원작사) → 구분
        if "원작사" in t:
            return "rs_원작사"
        if "테라핀" in t:
            return "rs_사"
        return None
    # 총매출 열 헤더는 "(요약본 정산기준매출)" 설명을 품으므로 정산기준매출보다 먼저 확정
    for kw, f in [("정산순매출", "정산순매출"), ("총매출", "총매출"),
                  ("정산기준매출", "정산기준매출"),
                  ("서비스월", "서비스월"), ("플랫폼", "플랫폼"), ("작품", "작품명"),
                  ("국가", "국가"), ("항목", "항목"), ("웹소설본부", "웹소설본부"),
                  ("외주작가", "rs대상"), ("런칭", "런칭일"), ("비고", "비고"),
                  ("구분", "구분")]:
        if kw in t:
            return f
    if t in ("NO", "NO."):
        return "no"
    return None


def _canon_etc(h):
    """기타수익(광고/IP) 헤더 1칸 → 표준 필드명."""
    t = _t(h)
    if "집행월" in t:
        return "집행월"
    if "서비스월" in t:
        return "서비스월"
    if "지급" in t and "월" in t:
        return "지급월"
    for kw, f in [("종류", "종류"), ("플랫폼", "플랫폼"), ("작품", "작품명"),
                  ("금액", "금액"), ("비고", "비고")]:
        if kw in t:
            return f
    if t in ("NO", "NO."):
        return "no"
    return None


# ── 공통 유틸 ───────────────────────────────────────────────
def parse_company(filename):
    return "수성" if "수성" in filename else "테라핀"


def parse_period(filename):
    """파일명에서 정산월(YYYY-MM) 추출. 'YYYY년_M월' 또는 'YYYY_MM' 모두 지원."""
    m = re.search(r"(20\d{2})\s*년?\D{0,3}(\d{1,2})\s*월", filename)
    if not m:
        m = re.search(r"(20\d{2})[._\-](\d{1,2})\b", filename)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    return None


def norm_month(v):
    """'2026.05월', datetime, '2026-5' 등 다양한 표기 → 'YYYY-MM'. 'N'/빈값 → None."""
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month"):
        return f"{v.year:04d}-{v.month:02d}"
    m = re.search(r"(20\d{2})\D{0,3}(\d{1,2})", str(v))
    return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}" if m else None


def add_month(period, n):
    """'YYYY-MM' + n개월. (정산서월 → 지급월 = +1)"""
    y, mo = map(int, period.split("-"))
    idx = (y * 12 + (mo - 1)) + n
    return f"{idx // 12:04d}-{idx % 12 + 1:02d}"


def _sheet_for_month(sheetnames, period):
    """정산월(YYYY-MM) → 매출리스트 월 시트명(예: '26.07')."""
    y, mo = period.split("-")
    cand = f"{int(y) % 100:02d}.{int(mo):02d}"
    if cand in sheetnames:
        return cand
    for s in sheetnames:                              # 공백/표기 흔들림 보정
        if _t(s) == _t(cand):
            return s
    return None


def _detect_header(ws, canon_fn, need=3, scan=12):
    """헤더 행 자동 인식: canon_fn 으로 매핑되는 칸이 need개 이상인 첫 행."""
    best_r, best_map, best_n = None, None, -1
    for r in range(1, scan + 1):
        cmap = {}
        for ci, cell in enumerate(ws[r], 1):
            f = canon_fn(cell.value)
            if f and f not in cmap:
                cmap[f] = ci
        if len(cmap) > best_n:
            best_r, best_map, best_n = r, cmap, len(cmap)
        if len(cmap) >= need and ("작품명" in cmap):
            return r, cmap
    return (best_r, best_map) if best_n >= need else (None, None)


def _num(v):
    """금액/율 → float. 콤마·공백·통화기호 제거. 빈값은 None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    try:
        return float(s) if s not in ("", "-", ".") else None
    except ValueError:
        return None


_SKIP = ("합계", "소계", "총계", "계", "small", "sum", "subtotal")


def _is_skip_row(values):
    joined = "".join(str(v) for v in values if v is not None)
    if joined.strip() == "":
        return True
    head = str(values[0] if values else "").strip()
    return any(k in head for k in _SKIP) or any(_t(joined) == _t(k) for k in _SKIP)


# ── 매출리스트 로더 ─────────────────────────────────────────
def load_sales_file(path, 정산서월, warnings):
    fn = os.path.basename(path)
    company = parse_company(fn)
    지급월 = add_month(정산서월, 1)                    # 정산서월 → 지급월(+1) = 시트월
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = _sheet_for_month(wb.sheetnames, 지급월)
    if not sheet:
        warnings.append(f"[매출] {fn}: {정산서월} 정산서용 지급월({지급월}) 시트 없음(시트예: {wb.sheetnames[:4]})")
        return []
    ws = wb[sheet]
    hr, cmap = _detect_header(ws, _canon_sales)
    if not cmap or "작품명" not in cmap:
        warnings.append(f"[매출] {fn}[{sheet}]: 헤더 인식 실패")
        return []
    recs = []
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if _is_skip_row(row):
            continue
        g = lambda f: (row[cmap[f] - 1] if f in cmap and cmap[f] - 1 < len(row) else None)
        작품 = g("작품명")
        순매출 = _num(g("정산순매출"))
        if (작품 in (None, "")) and 순매출 is None and _num(g("총매출")) is None:
            continue
        recs.append({
            "source": "매출", "회사": company, "정산서월": 정산서월, "지급월": 지급월, "시트": sheet,
            "항목": g("항목"), "국가": g("국가"), "구분": g("구분"),
            "플랫폼": g("플랫폼"), "작품명": 작품, "서비스월": g("서비스월"),
            "서비스월_norm": norm_month(g("서비스월")),
            "총매출": _num(g("총매출")), "rs_사": _num(g("rs_사")),
            "정산기준매출": _num(g("정산기준매출")), "rs_원작사": _num(g("rs_원작사")),
            "원작료": _num(g("원작료")), "정산순매출": 순매출,
            "웹소설본부": g("웹소설본부"), "rs대상": g("rs대상"), "비고": g("비고"),
        })
    return recs


# ── 기타수익 로더 (광고 + IP) ───────────────────────────────
def load_etc_file(path, 정산서월, warnings):
    fn = os.path.basename(path)
    company = parse_company(fn)
    지급월 = add_month(정산서월, 1)
    wb = load_workbook(path, read_only=True, data_only=True)
    recs = []
    for sheet, kind in [("정산리스트(광고)", "광고"), ("정산리스트(IP)", "IP")]:
        if sheet not in wb.sheetnames:
            if kind == "IP":
                continue                              # 수성은 IP 없음(정상)
            warnings.append(f"[기타] {fn}: '{sheet}' 시트 없음")
            continue
        ws = wb[sheet]
        hr, cmap = _detect_header(ws, _canon_etc)
        if not cmap or "금액" not in cmap:
            warnings.append(f"[기타] {fn}[{sheet}]: 헤더 인식 실패")
            continue
        # 지급월 컬럼이 3개(외부제작사/원작사/작가)로 분리 → 서브헤더(hr+1)에서 '원작사' 열을 찾는다
        col_원작사 = None
        sub = [c.value for c in ws[hr + 1]] if hr + 1 <= ws.max_row else []
        for idx, val in enumerate(sub, start=1):
            if val is not None and "원작사" in _t(val):
                col_원작사 = idx
                break
        if col_원작사 is None:                           # 폴백: 단일 지급월 열
            col_원작사 = cmap.get("지급월")
        for row in ws.iter_rows(min_row=hr + 2, values_only=True):
            if _is_skip_row(row):
                continue
            g = lambda f: (row[cmap[f] - 1] if f in cmap and cmap[f] - 1 < len(row) else None)
            원작사지급 = (row[col_원작사 - 1] if col_원작사 and col_원작사 - 1 < len(row) else None)
            금액 = _num(g("금액"))
            작품 = g("작품명")
            if 금액 is None and 작품 in (None, ""):
                continue
            recs.append({
                "source": "기타수익", "회사": company, "정산서월": 정산서월,
                "종류구분": kind,
                "기준월": g("집행월") if kind == "광고" else g("서비스월"),
                "기준월_norm": norm_month(g("집행월") if kind == "광고" else g("서비스월")),
                "종류": g("종류"), "플랫폼": g("플랫폼"), "작품명": 작품,
                "원작사지급월": 원작사지급, "지급월_norm": norm_month(원작사지급),
                "금액": 금액, "비고": g("비고"),
            })
    return recs


# ── 통합 진입점 ─────────────────────────────────────────────
def load_all(inbox_dir, 정산서월):
    """inbox/매출리스트 + inbox/기타수익 전체 로드 → 표준 레코드 + 적재 리포트.
       정산서월(예: 2026-06)을 받아 내부적으로 지급월(+1=2026-07) 시트를 사용."""
    지급월 = add_month(정산서월, 1)
    sales, etc, warnings, files = [], [], [], []
    for path in sorted(glob.glob(os.path.join(inbox_dir, "매출리스트", "*.xlsx"))):
        n0 = len(sales)
        sales += load_sales_file(path, 정산서월, warnings)
        files.append({"종류": "매출", "파일": os.path.basename(path),
                      "회사": parse_company(os.path.basename(path)), "행수": len(sales) - n0})
    for path in sorted(glob.glob(os.path.join(inbox_dir, "기타수익", "*.xlsx"))):
        n0 = len(etc)
        etc += load_etc_file(path, 정산서월, warnings)
        files.append({"종류": "기타수익", "파일": os.path.basename(path),
                      "회사": parse_company(os.path.basename(path)), "행수": len(etc) - n0})
    # 기타수익은 전기간 누적 원장 → 당월 반영분은 '원작사 정산월 = 정산서월'인 행
    # (매출은 지급월 시트 기준, 기타수익은 원작사 정산월=정산서월 기준 — 같은 서비스월 배치)
    당월광고 = sum(1 for r in etc if r["종류구분"] == "광고" and r["지급월_norm"] == 정산서월)
    당월IP = sum(1 for r in etc if r["종류구분"] == "IP" and r["지급월_norm"] == 정산서월)
    보류 = sum(1 for r in etc if r["지급월_norm"] is None)
    return {"정산서월": 정산서월, "지급월": 지급월, "매출": sales, "기타수익": etc,
            "files": files, "warnings": warnings,
            "기타수익_당월지급": {"광고": 당월광고, "IP": 당월IP, "지급월미정(N등)": 보류}}
