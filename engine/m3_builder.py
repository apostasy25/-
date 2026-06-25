# -*- coding: utf-8 -*-
"""
M3 정산서 빌더 — RAW(매출 + 기타수익 누적)에서 정산서 상세를 직접 재구성.

· 연재 상세  : 매출 파일(월시트). 정산금액 = ROUND(정산기준매출 × 원작사RS).
· 광고 상세  : 기타수익 누적('정산리스트(광고)'). 원작사정산월 기준.
              정산금액 = ROUND(공급가액 × 원작사RS).
· MG 작품    : 전체 이력(연재+광고) 누적 + MG총액행 + 잔여 MG.  잔여<0 → "MG 차감중".
· 非MG 작품  : 당월(지급월 시트) 분만.
· 누적상계는 선인세MG마스터 권위값과 교차검증(불일치 시 플래그).

직전 정산서(수기본)에 의존하지 않는다 — 휴먼 에러 차단.
"""
import re
import glob
import os
from openpyxl import load_workbook

import sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "gui", "master"))
import master_io as MIO


def _m2k(s):
    m = re.search(r"(\d{4})[.\-/]?\s*(\d{1,2})", str(s) or "")
    return f"{m.group(1)}-{int(m.group(2)):02d}" if m else ""


def _norm(s):
    return str(s).replace(" ", "").strip() if s else ""


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


# ── 매출 RAW ───────────────────────────────────────────────
def _sales_header(ws):
    for r in range(1, 9):
        row = [str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)]
        if any("작품명" in v for v in row):
            idx = {}
            for c, v in enumerate(row, 1):
                if "작품명" in v: idx["작품"] = c
                elif v.strip() == "항목": idx["항목"] = c
                elif "국가" in v: idx["국가"] = c
                elif "구분" in v: idx["구분"] = c
                elif "플랫폼" in v: idx["플랫폼"] = c
                elif "런칭" in v: idx["런칭"] = c
                elif "서비스월" in v: idx["서비스월"] = c
                elif "총매출" in v: idx["총매출"] = c
                elif "정산기준매출" in v: idx["정산기준"] = c
                elif "원작료" in v: idx["원작료"] = c
                elif "RS" in v and "원작사" in v: idx["RS"] = c
            return r, idx
    return None, {}


def load_sales(sales_path):
    """매출 전체를 1회 로드해 작품별 캐시 생성 → {norm작품: [row,...]}.
       row: 지급시트·국가·구분·플랫폼·런칭·서비스월·정산기준·원작료항목여부."""
    from collections import defaultdict
    wb = load_workbook(sales_path, data_only=True, read_only=True)
    cache = defaultdict(list)
    for sh in wb.sheetnames:
        if sh == "list":
            continue
        ws = wb[sh]
        hr, idx = _sales_header(ws)
        if not hr or "원작료" not in idx:
            continue
        for row in ws.iter_rows(min_row=hr + 1, values_only=True):
            if len(row) < idx["원작료"]:
                continue
            nm = _norm(row[idx["작품"] - 1])
            if not nm:
                continue
            cache[nm].append({
                "지급시트": sh,
                "국가": row[idx["국가"] - 1], "구분": row[idx["구분"] - 1],
                "플랫폼": row[idx["플랫폼"] - 1],
                "런칭": row[idx["런칭"] - 1] if idx.get("런칭") else None,
                "서비스월": _m2k(row[idx["서비스월"] - 1]) if idx.get("서비스월") else "",
                "총매출": _num(row[idx["총매출"] - 1]) if idx.get("총매출") else 0,
                "정산기준": _num(row[idx["정산기준"] - 1]),
            })
    return cache


def 연재상세(sales_path, 작품, rs, 서비스월_cutoff=None, 당월_지급시트=None, cache=None,
           서비스월_set=None, 지급시트_set=None, 누적_cutoff=None):
    """연재 상세행. 누적_cutoff 주면 서비스월 ≤ cutoff 전체이력(분기/반기/해외 누적, dedup)."""
    if cache is None:
        cache = load_sales(sales_path)
    rows = cache.get(_norm(작품), [])
    if 서비스월_set is not None or 누적_cutoff is not None:
        # (서비스월·국가·플랫폼·구분)별 최신 지급시트 1건만(정정분 중복 제거)
        best = {}
        for d in rows:
            svc = d["서비스월"]
            if 서비스월_set is not None and svc not in 서비스월_set:
                continue
            if 누적_cutoff is not None and (not svc or svc > 누적_cutoff):
                continue
            key = (svc, d["국가"], d["플랫폼"], d["구분"])
            if key not in best or _sheet_order(d["지급시트"]) > _sheet_order(best[key]["지급시트"]):
                best[key] = d
        sel = list(best.values())
    else:
        sel = []
        for d in rows:
            if 지급시트_set is not None:
                if d["지급시트"] not in 지급시트_set:
                    continue
            elif 당월_지급시트 and d["지급시트"] != 당월_지급시트:
                continue
            if 서비스월_cutoff and d["서비스월"] and d["서비스월"] > 서비스월_cutoff:
                continue
            sel.append(d)
    out = []
    for d in sel:
        Bv = d["정산기준"]
        out.append({
            "국가": d["국가"], "구분": d["구분"], "플랫폼": d["플랫폼"], "런칭": d["런칭"],
            "서비스월": d["서비스월"], "총매출": d["총매출"], "지급시트": d["지급시트"],
            "정산기준순매출": round(Bv), "RS": rs, "정산금액": round(round(Bv) * rs), "비고": "",
        })
    out.sort(key=lambda x: x["서비스월"])
    return out


# ── 기타수익(광고) 누적 RAW ─────────────────────────────────
def load_etc(etc_path):
    """기타수익(광고) 누적을 1회 로드 → {norm작품: [row,...]}.
       경로가 없거나 파일이 없으면 빈 캐시(기타수익 없음)로 처리."""
    from collections import defaultdict
    cache = defaultdict(list)
    if not etc_path or not os.path.exists(etc_path):
        return cache
    wb = load_workbook(etc_path, data_only=True, read_only=True)
    if "정산리스트(광고)" not in wb.sheetnames:
        return cache
    ws = wb["정산리스트(광고)"]
    for row in ws.iter_rows(min_row=8, values_only=True):
        if len(row) < 10:
            continue
        nm = _norm(row[5])
        if not nm:
            continue
        cache[nm].append({
            "종류": row[3], "플랫폼": row[4],
            "원작사정산월": _m2k(row[7]), "금액": _num(row[9]),
        })
    return cache


def 광고상세(etc_path, 작품, rs, 원작사정산월_cutoff, 당월only=None, cache=None,
           원작사정산월_set=None):
    """광고 상세행(원작사 정산 대상). cache 주면 캐시 사용. 원작사정산월_set(기간) 지원."""
    if cache is None:
        cache = load_etc(etc_path)
    rows = cache.get(_norm(작품), [])
    out = []
    for d in rows:
        wol = d["원작사정산월"]
        if not wol:
            continue
        if 원작사정산월_set is not None:
            if wol not in 원작사정산월_set:
                continue
        elif 당월only and wol != 당월only:
            continue
        if 원작사정산월_cutoff and wol > 원작사정산월_cutoff:
            continue
        금액 = d["금액"]
        out.append({
            "국가": "-", "구분": "광고수익", "플랫폼": d["플랫폼"],
            "런칭": "-", "서비스월": wol, "총매출": 0,
            "정산기준순매출": round(금액), "RS": rs, "정산금액": round(round(금액) * rs),
            "비고": d["종류"],
        })
    out.sort(key=lambda x: x["서비스월"])
    return out


# ── 작품 단위 상세(연재+광고, MG면 누적+MG행) ────────────────
def build_work(sales_path, etc_path, 작품, rs, 정산서월, is_mg, mg_total=0,
               mg_anchor=None):
    """반환: dict(rows, 정산금액합, is_mg, 잔여, 차감중, 검증)."""
    지급월 = _next_month(정산서월)              # 매출 지급월 시트명
    sheet = _ymd_to_sheet(지급월)
    if is_mg:
        연재 = 연재상세(sales_path, 작품, rs, 서비스월_cutoff=None)
        광고 = 광고상세(etc_path, 작품, rs, 원작사정산월_cutoff=정산서월)
    else:
        연재 = 연재상세(sales_path, 작품, rs, 당월_지급시트=sheet)
        광고 = 광고상세(etc_path, 작품, rs, 원작사정산월_cutoff=None, 당월only=정산서월)
    rows = 연재 + 광고
    누적상계 = sum(x["정산금액"] for x in rows)
    res = {"작품": 작품, "rows": rows, "연재행": len(연재), "광고행": len(광고),
           "정산금액합": 누적상계, "is_mg": is_mg}
    if is_mg:
        잔여 = 누적상계 - mg_total
        res.update({"mg_total": mg_total, "잔여": 잔여, "차감중": 잔여 < 0,
                    "공급가액": 0 if 잔여 < 0 else max(0, 잔여)})
        if mg_anchor is not None:
            res["검증"] = ("OK" if abs(누적상계 - mg_anchor) <= 10
                          else f"불일치 재구성{누적상계:,} vs 권위{mg_anchor:,}")
    else:
        res["공급가액"] = 누적상계
    return res


# ── 날짜 헬퍼 ──────────────────────────────────────────────
def _next_month(ym):
    y, m = map(int, ym.split("-"))
    return f"{y + 1}-01" if m == 12 else f"{y}-{m + 1:02d}"


def _ymd_to_sheet(ym):
    """'2026-06' → 매출 시트명 '26.06'(경계월 < > 는 호출측에서 보정)."""
    y, m = ym.split("-")
    return f"{y[2:]}.{m}"


# ── 환율(통화코드 기반 일반화) ──────────────────────────────
def load_fx(master_path):
    """환율마스터 → {(통화코드, 기준월): 환율}. JPY/USD/EUR 등 다통화."""
    from openpyxl import load_workbook as _lw
    ws = _lw(master_path, data_only=True)["환율마스터"]
    fx = {}
    for r in range(1, ws.max_row + 1):
        cur = ws.cell(r, 1).value
        if cur in ("JPY", "USD", "EUR", "CNY", "TWD"):
            try:
                fx[(str(cur), str(ws.cell(r, 2).value))] = float(ws.cell(r, 3).value)
            except (TypeError, ValueError):
                pass
    return fx


def fx_rate(fx, 통화, 기준월):
    """(통화, 기준월) 환율. 없으면 None(→ 호출측에서 '환율 미입력' 플래그)."""
    return fx.get((통화, 기준월))


def 통화of(플랫폼, 국가, 업체=None):
    """해외 정산 통화 자동판정. KRW 정산(환율 비대상)이면 None."""
    p, g = str(플랫폼 or ""), str(국가 or "")
    if 업체 in ("신죠샤", "도쿠마 쇼텐"):
        return "JPY"                       # 일본 사업자 직접 청구
    if "만타" in p:
        return "USD"                       # 북미/스페인 만타
    if "코미코" in p or g == "일본":
        return "JPY"
    return None                            # 그 외 해외는 매출 파일이 이미 KRW


def to_foreign(krw, fx, 통화, 기준월, ndigits=0):
    """KRW → 외화 환산 = KRW / 환율(통화,기준월). 환율 없으면 None."""
    rate = fx_rate(fx, 통화, 기준월)
    if not rate:
        return None
    return round(krw / rate, ndigits)


def _sheet_order(sheet):
    """지급시트명 → 정렬키. '26.07'→202607, '<25.12'→202512, '26.01>'→202601."""
    import re as _re
    m = _re.search(r"(\d{2})\.(\d{2})", str(sheet) or "")
    if not m:
        return 0
    return 2000 * 100 + int(m.group(1)) * 100 + int(m.group(2)) + 200000
