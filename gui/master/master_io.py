# -*- coding: utf-8 -*-
"""
마스터 입출력 엔진 — 마스터 관리 GUI의 토대.

담당 기능 (화면 코드와 분리된 순수 로직):
  1) 자동 백업      : 저장 시점마다 타임스탬프 사본 생성, 최신 N개만 롤링 보관
  2) 동시편집 잠금  : 원드라이브 공용 폴더에서 여러 명이 동시에 마스터를
                      편집하다 충돌·유실되는 것을 방지. 편집 시 .lock 생성,
                      다른 사용자는 읽기전용/경고. 비정상 종료 시 stale 자동 판정.
  3) 시트 읽기/저장 : 헤더 자동 인식, '예시' 행 제외, 다른 시트 보존하며 저장.

원칙: 잠금은 '편집'에만 건다(읽기끼리는 충돌 없음). 원드라이브 동기화 지연이
      있어 100% 원자적이지는 않으나, 소수 정산팀 환경에서 충돌을 사실상 제거한다.
"""
import os
import json
import shutil
import getpass
import socket
import datetime
from openpyxl import load_workbook


# ── 공통 ────────────────────────────────────────────────────
def whoami():
    """현재 사용자/PC 식별 (잠금 표시·변경 이력에 사용)."""
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    try:
        pc = socket.gethostname()
    except Exception:
        pc = "unknown"
    return user, pc


def _ts(fmt="%Y-%m-%d_%H%M"):
    return datetime.datetime.now().strftime(fmt)


# ── 1) 자동 백업 ─────────────────────────────────────────────
def backup_master(master_path, backup_dir=None, keep=30):
    """마스터 사본을 backup/ 에 타임스탬프로 저장하고 최신 keep개만 남긴다."""
    if not os.path.exists(master_path):
        return None
    base_dir = os.path.dirname(os.path.abspath(master_path))
    backup_dir = backup_dir or os.path.join(base_dir, "backup")
    os.makedirs(backup_dir, exist_ok=True)
    stem = os.path.splitext(os.path.basename(master_path))[0]
    dst = os.path.join(backup_dir, f"{stem}_{_ts()}.xlsx")
    # 같은 분/초에 여러 번 저장돼도 덮어쓰지 않도록 고유 접미사 보장
    if os.path.exists(dst):
        base = os.path.join(backup_dir, f"{stem}_{_ts('%Y-%m-%d_%H%M%S')}")
        dst, i = base + ".xlsx", 1
        while os.path.exists(dst):
            dst = f"{base}_{i}.xlsx"
            i += 1
    shutil.copy2(master_path, dst)
    _prune_backups(backup_dir, stem, keep)
    return dst


def _prune_backups(backup_dir, stem, keep):
    files = [os.path.join(backup_dir, f) for f in os.listdir(backup_dir)
             if f.startswith(stem + "_") and f.endswith(".xlsx")]
    files.sort(key=os.path.getmtime, reverse=True)   # 최신 우선
    for old in files[keep:]:
        try:
            os.remove(old)
        except OSError:
            pass
    return len(files)


# ── 2) 동시편집 잠금 ─────────────────────────────────────────
def _lock_path(master_path):
    return os.path.splitext(master_path)[0] + ".lock"


def read_lock(master_path):
    """현재 잠금 정보 dict 또는 None. (손상된 lock 파일은 None 취급)"""
    lp = _lock_path(master_path)
    if not os.path.exists(lp):
        return None
    try:
        with open(lp, "r", encoding="utf-8") as f:
            info = json.load(f)
        info["_age_sec"] = (datetime.datetime.now()
                            - datetime.datetime.fromisoformat(info["time"])).total_seconds()
        return info
    except Exception:
        return None


def is_stale(info, ttl_min=30):
    """heartbeat 갱신이 ttl_min 분 넘게 끊긴 잠금 = stale(비정상 종료 추정)."""
    return bool(info) and info.get("_age_sec", 0) > ttl_min * 60


def acquire_lock(master_path, ttl_min=30):
    """편집 잠금 시도.
       성공 → (True, 내정보) / 실패(타인이 사용 중) → (False, 보유자정보)."""
    user, pc = whoami()
    info = read_lock(master_path)
    if info and not is_stale(info, ttl_min):
        if not (info.get("user") == user and info.get("pc") == pc):
            return False, info               # 타인이 편집 중 → 읽기전용
    _write_lock(master_path, user, pc)       # 신규/내것/stale → 점유
    return True, read_lock(master_path)


def refresh_lock(master_path):
    """편집 중 주기적 heartbeat (예: 30초마다 호출) → stale 오판 방지."""
    info = read_lock(master_path)
    user, pc = whoami()
    if info and info.get("user") == user and info.get("pc") == pc:
        _write_lock(master_path, user, pc)
        return True
    return False


def release_lock(master_path):
    """내 잠금 해제(파일 삭제). 내것이 아니면 건드리지 않음."""
    info = read_lock(master_path)
    user, pc = whoami()
    if info and info.get("user") == user and info.get("pc") == pc:
        try:
            os.remove(_lock_path(master_path))
        except OSError:
            pass
        return True
    return False


def force_release_lock(master_path):
    """stale 잠금 강제 해제(사용자가 확인 후). 항상 삭제."""
    try:
        os.remove(_lock_path(master_path))
        return True
    except OSError:
        return False


def _write_lock(master_path, user, pc):
    payload = {"user": user, "pc": pc, "time": datetime.datetime.now().isoformat(timespec="seconds")}
    with open(_lock_path(master_path), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)


# ── 3) 시트 읽기/저장 ────────────────────────────────────────
def _header_row(ws, max_scan=8):
    """헤더 행 자동 인식: 셀이 가장 많이 채워진 상단 행."""
    best_r, best_n = 1, -1
    upper = min(max_scan, ws.max_row or 1)
    for r in range(1, upper + 1):
        n = sum(1 for c in ws[r] if c.value not in (None, ""))
        if n > best_n:
            best_r, best_n = r, n
    return best_r


def read_sheet(master_path, sheet, drop_example=True):
    """시트 → (headers, rows[list[dict]]). '예시' 표기 행은 기본 제외."""
    wb = load_workbook(master_path, data_only=True, read_only=True)
    ws = wb[sheet]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    if not grid:
        return [], []
    hr = _header_row(ws)
    headers = [str(v).strip() if v is not None else f"col{i+1}"
               for i, v in enumerate(grid[hr - 1])]
    rows = []
    for raw in grid[hr:]:
        cells = list(raw) + [None] * (len(headers) - len(raw))
        if all(v in (None, "") for v in cells):
            continue
        rowtext = " ".join(str(v) for v in cells if v is not None)
        if drop_example and "예시" in rowtext:
            continue
        rows.append({headers[i]: cells[i] for i in range(len(headers))})
    return headers, rows


def append_rows(master_path, sheet, rows, do_backup=True):
    """rows(list[dict])를 해당 시트 맨 아래에 추가(기존 행·서식·수식 모두 보존)."""
    if do_backup:
        backup_master(master_path)
    wb = load_workbook(master_path)
    ws = wb[sheet]
    hr = _header_row(ws)
    headers = [str(ws.cell(hr, c).value).strip() if ws.cell(hr, c).value is not None
               else f"col{c}" for c in range(1, ws.max_column + 1)]
    last = ws.max_row
    for i, row in enumerate(rows):
        for c, h in enumerate(headers, 1):
            if h in row:
                ws.cell(last + 1 + i, c).value = row[h]
    wb.save(master_path)
    return len(rows)


def save_sheet(master_path, sheet, headers, rows, do_backup=True):
    """rows(list[dict])를 해당 시트에 다시 기록. 다른 시트·서식은 보존.
       저장 전 자동 백업. (수식 문자열은 그대로 보존되며 Excel에서 재계산.)"""
    if do_backup:
        backup_master(master_path)
    wb = load_workbook(master_path)            # 서식·수식 보존
    ws = wb[sheet]
    hr = _header_row(ws)
    # 기존 데이터 행 비우기 (헤더 아래 전부)
    if ws.max_row > hr:
        ws.delete_rows(hr + 1, ws.max_row - hr)
    # 새 데이터 기록
    for i, row in enumerate(rows):
        for j, h in enumerate(headers):
            ws.cell(hr + 1 + i, j + 1).value = row.get(h)
    wb.save(master_path)
    return len(rows)


# ── 4) 엑셀 다운로드(내보내기) / 업로드(가져오기: 덮어쓰기·병합) ──────────
# 병합 시 행 동일성 판단 키(시트별). 미정의 시트는 병합 모드에서 덮어쓰기로 처리.
MERGE_KEYS = {
    "업체마스터": ("업체ID",),
    "작품Alias마스터": ("표준작품명",),
    "이메일마스터": ("업체ID",),
    "환율마스터": ("통화", "적용월"),
    "선인세MG마스터": ("업체ID", "작품명"),
    "이월마스터": ("업체ID", "작품명", "발생월"),
    "예외규칙마스터": ("표준작품명", "구간"),
}


def master_sheets(master_path):
    """마스터 파일의 시트 이름 목록(데이터 시트)."""
    wb = load_workbook(master_path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def export_master(master_path, dest_path):
    """현재 마스터 전체를 dest_path 로 복사(다운로드용 독립 사본)."""
    import shutil as _sh
    _sh.copy(master_path, dest_path)
    return dest_path


def _row_key(row, key_cols):
    return tuple(str(row.get(k, "")).strip() for k in key_cols)


def import_master(master_path, src_path, mode="overwrite", only_sheets=None,
                  do_backup=True):
    """업로드한 src_path(수정본)를 master_path 에 반영.
       mode='overwrite' : 공통 시트를 src 내용으로 통째 교체
       mode='merge'     : 키(MERGE_KEYS) 기준 upsert(기존 갱신 + 신규 추가)
       반환: {시트: {'mode':.., 'before':n, 'after':n, 'added':a, 'updated':u}}"""
    if do_backup:
        backup_master(master_path)
    tgt_sheets = set(master_sheets(master_path))
    src_sheets = set(master_sheets(src_path))
    sheets = [s for s in (only_sheets or sorted(tgt_sheets & src_sheets))
              if s in tgt_sheets and s in src_sheets]
    result = {}
    for sheet in sheets:
        s_headers, s_rows = read_sheet(src_path, sheet, drop_example=True)
        if not s_headers:
            continue
        m_headers, m_rows = read_sheet(master_path, sheet, drop_example=False)
        headers = m_headers or s_headers
        if mode == "merge" and sheet in MERGE_KEYS:
            keys = MERGE_KEYS[sheet]
            idx = {_row_key(r, keys): i for i, r in enumerate(m_rows)}
            added = updated = 0
            for sr in s_rows:
                k = _row_key(sr, keys)
                if k in idx:
                    m_rows[idx[k]].update(sr)
                    updated += 1
                else:
                    m_rows.append(sr)
                    idx[k] = len(m_rows) - 1
                    added += 1
            save_sheet(master_path, sheet, headers, m_rows, do_backup=False)
            result[sheet] = {"mode": "merge", "before": len(m_rows) - added,
                             "after": len(m_rows), "added": added, "updated": updated}
        else:
            real = "overwrite" if mode == "overwrite" else "overwrite(키 미정의→덮어쓰기)"
            save_sheet(master_path, sheet, headers, s_rows, do_backup=False)
            result[sheet] = {"mode": real, "before": len(m_rows),
                             "after": len(s_rows), "added": 0, "updated": 0}
    return result


def vendor_id_name_map(master_path):
    """업체ID → '업체ID | 업체명' 표시 문자열 매핑(가시성 강화용)."""
    try:
        _h, rows = read_sheet(master_path, "업체마스터", drop_example=True)
    except Exception:                              # noqa: BLE001
        return {}
    out = {}
    for r in rows:
        vid = str(r.get("업체ID", "")).strip()
        if vid:
            out[vid] = f"{vid} | {r.get('업체명', '')}"
    return out


def vendor_label(master_path, vid):
    """단일 업체ID → '업체ID | 업체명'. 없으면 ID 그대로."""
    return vendor_id_name_map(master_path).get(str(vid).strip(), str(vid))
