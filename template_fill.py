# -*- coding: utf-8 -*-
"""정산서_양식_기초.xlsx 템플릿을 복사해 데이터만 채우는 엔진.

원칙(작가정산_원작료 합의): 셀 병합·배경색·글꼴·테두리·날짜형식·행높이·열너비를
절대 변경하지 않는다. 템플릿 시트를 복제해 값/수식만 채운다.

유형별 템플릿 시트:
  사업자       → 테라핀_사업자
  사업자(환율) → 테라핀_사업자(환율)
  개인         → 테라핀_개인
  해외         → 테라핀_해외
  수성         → 수성_기초

이 모듈은 사업자 유형을 먼저 구현한다. 섹션1(작품 1개분)을 '스타일 원본'으로 삼아
작품 수만큼 상세 섹션을 복제하고, 증빙표 품목행도 작품 수만큼 채운다. 수식
(SUMIFS·COUNTIF·ROUND·SUM)은 템플릿 그대로 유지되어 자동 계산된다.
"""
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, PatternFill
from openpyxl.styles import PatternFill, Border

TEMPLATE_SHEET = {
    "사업자": "테라핀_사업자",
    "사업자_환율": "테라핀_사업자(환율)",
    "개인": "테라핀_개인",
    "해외": "테라핀_해외",
    "수성": "수성_기초",
}


def _copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.number_format = src.number_format


def open_single_sheet(template_path, type_key):
    """템플릿 워크북을 열고 해당 유형 시트만 남긴다(서식 보존)."""
    wb = load_workbook(template_path)
    keep = TEMPLATE_SHEET[type_key]
    if keep not in wb.sheetnames:
        raise KeyError(f"템플릿에 '{keep}' 시트가 없습니다: {template_path}")
    for sn in list(wb.sheetnames):
        if sn != keep:
            del wb[sn]
    return wb, wb[keep]


# 사업자 템플릿의 고정 좌표(분석 결과)
PROOF_FIRST = 6          # 증빙표 품목 첫 행
PROOF_LAST = 12          # 증빙표 품목 마지막 슬롯
PROOF_SUM = 13           # 합계행
SEC1_NUM = 15            # 상세 섹션1 자동번호행
SEC1_HDR = 16            # 헤더행
SEC1_DATA0 = 17          # 데이터 첫 행
SEC1_DATA_N = 5          # 템플릿 섹션당 데이터 슬롯 수
SEC1_SUB = 22            # 소계행
MG_TOTAL_ROW = 44        # MG 총액 행(템플릿)
MG_DATA0 = 45            # MG 데이터 첫 행
MG_REMAIN_ROW = 53       # 잔여 MG 행
DET_COLS = ["국가", "구분", "플랫폼", "작품", "런칭", "서비스월", "총매출", "정산기준순매출", "RS"]
# 상세 열 위치: B순번 C국가 D구분 E플랫폼 F작품 G런칭 H서비스월 I총매출 J정산기준순매출 K원작사RS L정산금액 M비고
DET_COL_IDX = {"국가": 3, "구분": 4, "플랫폼": 5, "작품": 6, "런칭": 7,
               "서비스월": 8, "총매출": 9, "정산기준순매출": 10, "RS": 11}


def _grab_row_style(ws, row, c0=2, c1=14):
    return [(c, ws.cell(row, c)._style) for c in range(c0, c1 + 1)]


def _svc_month_value(v):
    """서비스월 'YYYY-MM' 문자열을 날짜로 변환(템플릿의 yyyy"년" mm"월" 형식 적용되도록).
       범위·기타 표기('2023.08월 ~ …', '2023년 08월')는 원본 문자열 유지."""
    import re
    import datetime as _dt
    if isinstance(v, str):
        m = re.fullmatch(r"\s*(\d{4})-(\d{1,2})\s*", v)
        if m:
            return _dt.datetime(int(m.group(1)), int(m.group(2)), 1)
    return v


def _section_merges(ws, top, bot):
    """top~bot 범위 내 병합을, 섹션 시작 기준 상대 오프셋으로 추출."""
    out = []
    for m in ws.merged_cells.ranges:
        if m.min_row >= top and m.max_row <= bot:
            out.append((m.min_row - top, m.min_col, m.max_row - top, m.max_col))
    return out


def fill_business(template_path, out_path, ctx):
    """사업자(테라핀_사업자) 템플릿에 데이터를 채워 out_path 로 저장.
    ctx = {year, month, vendor, email, 마감, 지급,
           works:[{작품, rows:[{국가,구분,플랫폼,작품,런칭,서비스월,총매출,정산기준순매출,RS}, ...]}, ...]}"""
    type_key = ctx.get("type_key", "사업자")
    wb, ws = open_single_sheet(template_path, type_key)
    yy, mm = ctx["year"], ctx["month"]

    # 1) 제목
    title = str(ws.cell(2, 2).value or "")
    title = title.replace("YYYY", str(yy)).replace("MM", f"{int(mm):02d}") \
                 .replace("[업체명]", ctx["vendor"])
    ws.cell(2, 2, title)

    works = ctx["works"]
    N = len(works)
    proof_slots = max(1, N)                 # 증빙표 품목 슬롯 = 실제 작품 수
    delta = proof_slots - 7                 # 상세 섹션 시작 이동량(음수=축소)
    PF = PROOF_FIRST                        # 6
    PL = PF + proof_slots - 1               # 증빙표 마지막 품목행
    PS = PF + proof_slots                   # 합계행 = 6+slots
    S_NUM = SEC1_NUM + delta                # 상세 섹션1 시작(이동)

    # 2) 상세 섹션 스타일 원본 추출(섹션1) + 증빙표 품목/합계 스타일
    num_style = _grab_row_style(ws, SEC1_NUM)
    hdr_style = _grab_row_style(ws, SEC1_HDR)
    data_style = _grab_row_style(ws, SEC1_DATA0)
    sub_style = _grab_row_style(ws, SEC1_SUB)
    mg_total_style = _grab_row_style(ws, MG_TOTAL_ROW)
    mg_data_style = _grab_row_style(ws, MG_DATA0)
    remain_style = _grab_row_style(ws, MG_REMAIN_ROW)
    mg_gray_fill = copy(ws.cell(MG_DATA0, 3).fill)        # MG 이전달 누적행 회색(복사본으로 고정)
    proof_item_style = _grab_row_style(ws, PROOF_FIRST + 1)   # 품목행(중간) 스타일
    proof_sum_style = _grab_row_style(ws, PROOF_SUM)          # 합계행 스타일
    proof_note = ws.cell(PROOF_SUM, 10).value                 # 합계행 이메일칸 안내문
    tpl_sum_fills = {c: (copy(ws.cell(PROOF_SUM, c).fill), copy(ws.cell(PROOF_SUM, c).border))
                     for c in range(2, 15)}                                       # 합계행 배경+테두리
    tpl_item_fills = {c: (copy(ws.cell(PROOF_FIRST + 1, c).fill), copy(ws.cell(PROOF_FIRST + 1, c).border))
                      for c in range(2, 15)}                                      # 품목행 배경+테두리
    proof_fix = []          # 저장 2단계에서 배경/테두리 재적용할 (행, {열:(fill,border)})
    proof_h = ws.row_dimensions[PROOF_FIRST].height
    hdr_vals = [ws.cell(SEC1_HDR, c).value for c in range(2, 15)]
    rowh = {r: ws.row_dimensions[r].height for r in (SEC1_NUM, SEC1_HDR, SEC1_DATA0, SEC1_SUB)}

    def _fb(row):                       # 행의 (fill,border) 맵 — 2단계 복원용
        return {c: (copy(ws.cell(row, c).fill), copy(ws.cell(row, c).border)) for c in range(2, 15)}
    fb_num, fb_hdr, fb_data, fb_sub = _fb(SEC1_NUM), _fb(SEC1_HDR), _fb(SEC1_DATA0), _fb(SEC1_SUB)
    fb_mgtot, fb_mgdata, fb_remain = _fb(MG_TOTAL_ROW), _fb(MG_DATA0), _fb(MG_REMAIN_ROW)
    detail_fix = []                     # (행, fb맵) — 재배치(delta≠0) 시 상세 행 서식 복원

    def _place_sum_row():
        """합계행을 PS 위치에 재구성(스타일·병합·SUM·안내문)."""
        for c, st in proof_sum_style:
            ws.cell(PS, c)._style = st
        ws.merge_cells(start_row=PS, start_column=2, end_row=PS, end_column=6)
        ws.merge_cells(start_row=PS, start_column=10, end_row=PS, end_column=14)
        ws.cell(PS, 2, "합계")
        ws.cell(PS, 7, f"=SUM(G{PF}:G{PL})")
        ws.cell(PS, 8, f"=SUM(H{PF}:H{PL})")
        ws.cell(PS, 9, f"=SUM(I{PF}:I{PL})")
        if proof_note:
            ws.cell(PS, 10, proof_note)

    def _remerge_proof_verticals():
        """증빙표 세로 병합(증빙서류·작성일자·이메일·마감·지급)을 N행에 맞춰 재설정."""
        ws.merge_cells(start_row=6, start_column=2, end_row=PL, end_column=3)
        ws.merge_cells(start_row=6, start_column=4, end_row=PL, end_column=4)
        ws.merge_cells(start_row=6, start_column=10, end_row=PL, end_column=12)
        ws.merge_cells(start_row=6, start_column=13, end_row=PL, end_column=13)
        ws.merge_cells(start_row=6, start_column=14, end_row=PL, end_column=14)

    # 3) 상세/합계 영역 초기화: 기존 섹션 제거 후 재구성
    last = ws.max_row
    if delta > 0:
        clear_start = PROOF_SUM            # 확장: 옛 합계행(13)부터
    elif delta < 0:
        clear_start = PS                   # 축소: 새 합계행 위치부터(남는 품목칸·옛 합계 제거)
    else:
        clear_start = SEC1_NUM
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= clear_start:
            ws.unmerge_cells(str(m))
    if delta != 0:                         # 증빙표 세로 병합 해제(재설정 위해)
        for mref in ["B6:C12", "D6:D12", "J6:L12", "M6:M12", "N6:N12"]:
            if mref in {str(m) for m in ws.merged_cells.ranges}:
                ws.unmerge_cells(mref)
    for r in range(clear_start, last + 1):
        for c in range(2, 15):
            ws.cell(r, c).value = None

    # 2-b) 증빙표 크기 조정
    if delta > 0:
        # 확장: 품목 행 추가
        for k in range(7, proof_slots):
            rr = PF + k
            for c, st in proof_item_style:
                ws.cell(rr, c)._style = st
            ws.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=6)
            if proof_h:
                ws.row_dimensions[rr].height = proof_h
            proof_fix.append((rr, tpl_item_fills))      # 새 품목행 크림 배경 재적용
        _remerge_proof_verticals()
        _place_sum_row()
        proof_fix.append((PS, tpl_sum_fills))           # 이동된 합계행 남색 재적용
    elif delta < 0:
        # 축소: 남는 품목칸(6+N..12)·옛 합계행(13) 스타일/병합 제거 → 투명화
        for rr in range(PS, PROOF_SUM + 1):
            for c in range(2, 15):
                cell = ws.cell(rr, c)
                cell.value = None
                cell.border = Border()
                cell.fill = PatternFill()
            ws.row_dimensions[rr].height = None
        _remerge_proof_verticals()
        _place_sum_row()
        proof_fix.append((PS, tpl_sum_fills))           # 이동된 합계행 남색 재적용

    sec_merges = _section_merges(load_workbook(template_path)[TEMPLATE_SHEET[type_key]],
                                 SEC1_NUM, SEC1_SUB)

    # 4) 작품별 섹션 재구성
    mg_remain_cell = {}      # 작품 → 잔여 MG 셀 주소
    mg_now_sum = {}          # 작품 → 당월 발생분 L셀 합(수식 문자열)
    mg_shade_rows = []       # (행, 당월여부) — 저장 직전 음영 일괄 적용
    r = S_NUM
    for w in works:
        is_mg = w.get("mg_total") is not None
        top = r
        # 자동번호행 — ② 현재 행 기준 동적 COUNTIF(섹션마다 갱신)
        for c, st in num_style:
            ws.cell(r, c)._style = st
        ws.row_dimensions[r].height = rowh[SEC1_NUM]
        ws.cell(r, 2, f'=COUNTIF($B$1:B{r-1},"*. 웹툰*")+1&". 웹툰 <{w["작품"]}> 원작료 정산 상세내역"')
        if delta:
            detail_fix.append((r, fb_num))
        r += 1
        # 헤더행
        hdr_row = r
        for i, (c, st) in enumerate(hdr_style):
            ws.cell(r, c)._style = st
            ws.cell(r, c, hdr_vals[i])
        ws.row_dimensions[r].height = rowh[SEC1_HDR]
        if delta:
            detail_fix.append((r, fb_hdr))
        r += 1
        mg_total_row = None
        if is_mg:                                            # MG 총액 행(데이터 위)
            for c, st in mg_total_style:
                ws.cell(r, c)._style = st
            ws.cell(r, 2, 0)
            ws.cell(r, 3, "웹툰화 계약시 원작자 수익분배금 MG")
            ws.cell(r, 12, w["mg_total"])
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
            ws.merge_cells(start_row=r, start_column=13, end_row=r, end_column=14)
            mg_total_row = r
            if delta:
                detail_fix.append((r, fb_mgtot))
            r += 1
        # 데이터행
        dstyle = mg_data_style if is_mg else data_style
        fb_d = fb_mgdata if is_mg else fb_data
        data_top = r
        now_cells = []                                       # MG 당월 발생행 L셀(증분 지급 계산용)
        for d in w["rows"]:
            for c, st in dstyle:
                ws.cell(r, c)._style = st
            ws.row_dimensions[r].height = rowh[SEC1_DATA0]
            ws.cell(r, 2, r - data_top + 1)                  # 순번
            for key, ci in DET_COL_IDX.items():
                val = _svc_month_value(d.get(key)) if key == "서비스월" else d.get(key)
                ws.cell(r, ci, val)
            ws.cell(r, 7).number_format = "yyyy-mm-dd"       # ⑤ 런칭일 YYYY-MM-DD 고정
            ws.cell(r, 12, f"=ROUND(J{r}*K{r},0)")           # 정산금액 (RS율·형식 템플릿 보존)
            if d.get("비고"):
                ws.cell(r, 13, d["비고"])
            # ⑥ MG 작품: 음영 적용은 저장 직전 일괄(이전달 회색 / 당월 무색)
            if is_mg:
                mg_shade_rows.append((r, bool(d.get("당월"))))
                if d.get("당월"):
                    now_cells.append(f"L{r}")
            if delta:
                detail_fix.append((r, fb_d))
            r += 1
        data_bot = r - 1
        # 소계 또는 잔여 MG 행
        if is_mg:
            for c, st in remain_style:
                ws.cell(r, c)._style = st
            ws.cell(r, 2, "합계")
            ws.cell(r, 12, f"=SUM(L{data_top}:L{data_bot})-L{mg_total_row}")
            ws.cell(r, 13, f'=IF(L{r}<0,"MG 차감중","")')
            mg_remain_cell[w["작품"]] = f"L{r}"
            mg_now_sum[w["작품"]] = "+".join(now_cells) if now_cells else "0"
            if delta:
                detail_fix.append((r, fb_remain))
        else:
            for c, st in sub_style:
                ws.cell(r, c)._style = st
            ws.cell(r, 2, "합계")
            ws.cell(r, 12, f"=SUM(L{data_top}:L{data_bot})")
            if delta:
                detail_fix.append((r, fb_sub))
        ws.row_dimensions[r].height = rowh[SEC1_SUB]
        sub_row = r
        # 병합: ④ 헤더 비고(M:N) + 데이터 비고(M:N) + 번호행(B:N) + 소계/잔여(B:K, M:N)
        ws.merge_cells(start_row=hdr_row, start_column=13, end_row=hdr_row, end_column=14)
        for rr in range(data_top, data_bot + 1):
            ws.merge_cells(start_row=rr, start_column=13, end_row=rr, end_column=14)
        ws.merge_cells(start_row=top, start_column=2, end_row=top, end_column=14)
        ws.merge_cells(start_row=sub_row, start_column=2, end_row=sub_row, end_column=11)
        ws.merge_cells(start_row=sub_row, start_column=13, end_row=sub_row, end_column=14)
        # ③ 섹션 사이 깨끗한 공백 행(테두리·배경·수식 초기화)
        gap_row = sub_row + 1
        for c in range(2, 15):
            gc = ws.cell(gap_row, c)
            gc.value = None
            gc.border = Border()
            gc.fill = PatternFill(fill_type=None)
        ws.row_dimensions[gap_row].height = None
        r = gap_row + 1

    # 4-b) 마지막 섹션 이후 잔여 템플릿 행 정리(테두리·배경 제거)
    for rr in range(r, last + 1):
        for c in range(2, 15):
            gc = ws.cell(rr, c)
            gc.value = None
            gc.border = Border()
            gc.fill = PatternFill(fill_type=None)

    # 5) 증빙표 품목행: 작품 수만큼 채우고 나머지 비움
    email = ctx.get("email") or "정산팀\nterapin_toonbill@terapinstudios.co.kr"
    품목라벨 = ctx.get("품목_label") or f"{int(mm):02d}월"
    held_set = {w["작품"] for w in works if w.get("held")}
    for i in range(PF, PL + 1):
        idx = i - PF
        if idx < len(works):
            wnm = works[idx]['작품']
            ws.cell(i, 5, f"원작료_{품목라벨} 정산 <{wnm}>")
            if wnm in mg_remain_cell:                        # MG 작품: 회수 후 당월 증분만 지급
                rc = mg_remain_cell[wnm]
                now = mg_now_sum.get(wnm, "0")
                # 공급가 = MAX(0,잔여) - MAX(0,잔여-당월분) → 차감중 0, 회수월 초과분, 이후 당월분만
                ws.cell(i, 7, f"=MAX(0,{rc})-MAX(0,{rc}-({now}))")
                ws.cell(i, 8, f"=ROUND(G{i}*10%,0)")
                ws.cell(i, 9, f'=IF({rc}<0,"MG 차감중",SUM(G{i}:H{i}))')
            else:
                # TEXTBEFORE/TEXTAFTER(엑셀 2021+ 전용) 의존 제거 → 작품명 직접 참조
                ws.cell(i, 7, f'=SUMIFS($L:$L,$F:$F,"{wnm}")')
                ws.cell(i, 8, f"=ROUND(G{i}*10%,0)")
                if wnm in held_set:                          # 증빙 미수취 → 지급 보류
                    ws.cell(i, 9, "증빙 미수취-보류")
                else:
                    ws.cell(i, 9, f"=G{i}+H{i}")
        else:
            ws.cell(i, 5).value = None
            for c in (7, 8, 9):
                ws.cell(i, c).value = None
    if works:
        ws.cell(PF, 10, email)
        # ① 증빙서류(B6:C 병합) · 작성일자(D 병합) 주입 — 날짜는 yyyy-mm-dd
        ws.cell(PF, 2, ctx.get("증빙구분", "세금계산서"))
        작성 = ctx.get("작성일자")
        if 작성 is not None:
            cell = ws.cell(PF, 4, 작성)
            if not isinstance(작성, str):
                cell.number_format = "yyyy-mm-dd"
        for col, key in ((13, "마감"), (14, "지급")):
            v = ctx.get(key)
            if v is not None:
                cell = ws.cell(PF, col, v)
                if not isinstance(v, str):
                    cell.number_format = "yyyy-mm-dd"

    # 6) 1차 저장 후, 재로드하여 배경(MG 음영·이동된 합계/품목) 재적용
    #    (빌드 중 워크북은 스타일 인덱스 상태 때문에 fill 변경이 반영되지 않음)
    wb.save(out_path)
    if mg_shade_rows or proof_fix or detail_fix:
        wb2 = load_workbook(out_path)
        ws2 = wb2.active
        # (1) 재배치된 상세/합계/품목 행의 배경+테두리 복원
        for (rr, fillmap) in detail_fix + proof_fix:
            for c, (fl, bd) in fillmap.items():
                ws2.cell(rr, c).fill = copy(fl)
                ws2.cell(rr, c).border = copy(bd)
        # (2) MG 음영(데이터 행 배경만 덮어씀 — 테두리는 위에서 복원된 것 유지)
        gray = PatternFill(patternType="solid", start_color="FFD9D9D9", end_color="FFD9D9D9")
        for (rr, is_now) in mg_shade_rows:
            f_ = PatternFill(fill_type=None) if is_now else gray
            for c in range(2, 15):
                ws2.cell(rr, c).fill = copy(f_)
        wb2.save(out_path)
    return out_path


# ════════════════════════════════════════════════════════════════
# 개인(작가) 정산서 — 테라핀_개인 템플릿 채움
#   증빙표(원천징수) + 1.웹툰 상세 + 2.기타 상세. 안내·참고사항 없음(수기 반영).
# ════════════════════════════════════════════════════════════════
def _capture_row_style(ws, row, cols):
    """행의 컬럼별 스타일·번호형식 캡처(복제용)."""
    return {c: (copy(ws.cell(row, c).font), copy(ws.cell(row, c).fill),
                copy(ws.cell(row, c).border), copy(ws.cell(row, c).alignment),
                ws.cell(row, c).number_format) for c in cols}


def _apply_row_style(ws, row, cap, height=None):
    for c, (f, fl, bd, al, nf) in cap.items():
        cl = ws.cell(row, c)
        cl.font = copy(f); cl.fill = copy(fl); cl.border = copy(bd)
        cl.alignment = copy(al); cl.number_format = nf
    if height:
        ws.row_dimensions[row].height = height


def _resize_block(ws, data0, tpl_n, new_n, sum_row):
    """[data0 .. data0+tpl_n-1] 데이터 블록을 new_n 행으로 조정.
       sum_row(합계행)·이하 내용은 insert/delete 로 자동 이동. 반환: 새 합계행 위치."""
    delta = new_n - tpl_n
    if delta > 0:
        ws.insert_rows(sum_row, delta)        # 합계행 앞에 delta 행 삽입
    elif delta < 0:
        ws.delete_rows(data0 + new_n, -delta)  # 초과 데이터행 삭제
    return sum_row + delta


def fill_personal(template_path, out_path, ctx):
    """테라핀_개인 템플릿 채움.
       ctx: year, month, author_disp, email, 작품명, 웹툰[], 기타[]
       각 row: 국가,구분,플랫폼,작품,런칭,서비스월,정산기준순매출,RS,비고"""
    wb, ws = open_single_sheet(template_path, "개인")
    yy, mm = ctx["year"], ctx["month"]
    웹툰 = ctx.get("웹툰") or []
    기타 = ctx.get("기타") or []
    DCOLS = list(range(2, 13))                 # B..L

    # 템플릿 고정 좌표: 웹툰 헤더9 데이터10~14(5) 합계15 / 기타 헤더18 데이터19~20(2) 합계21
    W_HDR, W_D0, W_TPL, W_SUM = 9, 10, 5, 15
    E_TPL = 2
    Nw = max(1, len(웹툰)); Ne = max(1, len(기타))

    # 데이터행 스타일 캡처(자료가 비어도 1행은 유지)
    w_style = _capture_row_style(ws, W_D0, DCOLS)
    e_style = _capture_row_style(ws, 19, DCOLS)
    rowh = ws.row_dimensions[W_D0].height or 20.2

    # 합계 병합(B:J)은 insert/delete 시 자동 이동 안 됨 → 해제 후 재병합
    for mc in ("B15:J15", "B21:J21"):
        try:
            ws.unmerge_cells(mc)
        except Exception:
            pass

    # 1) 웹툰 블록 조정 → 합계행 이동
    w_sum = _resize_block(ws, W_D0, W_TPL, Nw, W_SUM)
    shift = Nw - W_TPL
    # 기타 좌표(웹툰 조정 후 이동)
    E_HDR = 18 + shift; E_D0 = 19 + shift; E_SUM = 21 + shift
    e_sum = _resize_block(ws, E_D0, E_TPL, Ne, E_SUM)
    # 새 합계행 위치에 B:J 병합 재설정
    ws.merge_cells(start_row=w_sum, start_column=2, end_row=w_sum, end_column=10)
    ws.merge_cells(start_row=e_sum, start_column=2, end_row=e_sum, end_column=10)

    # 2) 스타일 재적용 + 데이터 채움
    def _fill_rows(rows, d0, cap, work_key="작품"):
        for i in range(max(len(rows), 1)):
            r = d0 + i
            _apply_row_style(ws, r, cap, rowh)
            if i < len(rows):
                d = rows[i]
                ws.cell(r, 2, i + 1)
                ws.cell(r, 3, d.get("국가"))
                ws.cell(r, 4, d.get("구분"))
                ws.cell(r, 5, d.get("플랫폼"))
                ws.cell(r, 6, d.get(work_key) or d.get("작품"))
                ws.cell(r, 7, d.get("런칭") or "-")
                ws.cell(r, 8, _svc_month_value(d.get("서비스월")))
                ws.cell(r, 9, d.get("정산기준순매출") or 0)
                ws.cell(r, 10, d.get("RS"))
                ws.cell(r, 10).number_format = "0.0%"
                ws.cell(r, 11, f"=ROUND(I{r}*J{r},0)")
                if d.get("비고"):
                    ws.cell(r, 12, d["비고"])
    _fill_rows(웹툰, W_D0, w_style)
    _fill_rows(기타, E_D0, e_style, work_key="작품")

    # 3) 합계·증빙 수식 갱신
    ws.cell(w_sum, 11, f"=SUM(K{W_D0}:K{w_sum-1})")
    ws.cell(e_sum, 11, f"=SUM(K{E_D0}:K{e_sum-1})")
    ws.cell(6, 9, f"=K{w_sum}+K{e_sum}")        # I6 총정산금액
    # J6/K6/L6(소득세/지방/실지급)·B6 이메일은 템플릿 수식 유지

    # 4) 헤더 텍스트
    ws.cell(2, 2, f"{yy}년 {mm}월 웹툰 원작료 정산서")
    ws.cell(6, 5, f"{ctx.get('author_disp','')}\n{ctx.get('email','')}")
    ws.cell(6, 6, ctx.get("작품명") or (웹툰[0].get("작품") if 웹툰 else ""))
    ws.title = f"{yy}-{mm}"
    wb.save(out_path)
    return out_path
