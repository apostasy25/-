# -*- coding: utf-8 -*-
"""
원작료 정산서 자동화 — 메인 오케스트레이터

사용법:
  python pipeline.py --company 테라핀 --year 2026 --month 5 --stage settle
  python pipeline.py --company 테라핀 --year 2026 --month 5 --stage pdf
  python pipeline.py --company 테라핀 --year 2026 --month 5 --stage mail
  python pipeline.py --company 테라핀 --year 2026 --month 5 --stage all   (검수 게이트 사이에 멈춤)

단계:
  settle : INBOX의 직전 발행 정산서 사본 → 정산서 생성 (output/정산서/)
  pdf    : 정산서 → 정산시트만 PDF → 비밀번호 암호화 (output/PDF/)
  mail   : 이메일 마스터 → .eml 초안 생성 (output/메일Draft/)

검수 게이트: 각 단계 후 manifest 로그를 남기고 멈춤. 사람이 검수 후 다음 단계 실행.
무발송: mail 단계는 .eml 까지만. 자동 발송 없음.
"""
import os, re, glob, sys, argparse, datetime, csv
from openpyxl import load_workbook

import config
from daterules import vendor_type, settle_dates, settle_dates_quarter
import settlement
import securepdf
import maildraft


# ──────────────────────────────────────────────────────────
def _vendor_from_filename(fn):
    """원본 파일명 → (회사, 업체명)."""
    b = os.path.basename(fn)
    company = "수성웹툰" if b.startswith("수성웹툰") else "테라핀"
    name = re.sub(r"^(테라핀|수성웹툰)__?\d{4}_\d{2}월_원작료_정산서_", "", b)
    name = name.replace(".xlsx", "").replace("_", " ").strip().strip("()_ ")
    return company, name


def group_of(company, name):
    """회사 + 원작사명 → 5개 정산 그룹 키 판정."""
    if company == "수성웹툰":
        return "수성"
    if any(k in name for k in config.OVERSEAS):       # 신죠샤·도쿠마 = 해외
        return "테라핀_해외"
    vt = vendor_type(name)
    if vt == "naver":
        return "테라핀_네이버"
    if vt == "ridi":
        return "테라핀_리디"
    return "테라핀_그외"                               # 카카오·개인·기타 사업자 등


def _name_from_settle_file(path, y, mo):
    """정산서/PDF 파일명 → 원작사명 추출 (그룹 판정용)."""
    b = os.path.basename(path)
    b = re.sub(r"^(원작료정산서_|수성웹툰_원작료정산서_)", "", b)
    b = re.sub(rf"_{y}{mo:02d}\.(xlsx|pdf)$", "", b)
    return b.replace("_", " ").strip()


def _outbase(company, y, mo):
    """산출물 루트 = output/YYYY-MM/회사.  월·회사별로 분리 보관(재생성·수정본 안전)."""
    comp = "수성" if company == "수성웹툰" else "테라핀"
    return os.path.join(config.OUTPUT, f"{y}-{mo:02d}", comp)


def _load_type_map(master_path):
    """업체명(정규화) → 유형(사업자/개인).  개인 판정의 단일 근거."""
    m = load_workbook(master_path, data_only=True)
    ws = m["업체마스터"]
    tm = {}
    for r in range(4, ws.max_row + 1):
        nm = ws.cell(r, 2).value
        typ = ws.cell(r, 4).value
        if nm and "예시" not in str(nm) and typ:
            tm[_norm(nm)] = str(typ).strip()
    return tm


def date_type(company, name, type_map):
    """날짜·양식 유형 판정.  개인 여부는 '이름'이 아니라 업체마스터 '유형' 기준."""
    t = type_map.get(_norm(name))
    if t is None:                                   # 정규화 부분일치 보정
        for k, v in type_map.items():
            if k and (k in _norm(name) or _norm(name) in k):
                t = v
                break
    if t == "개인":
        return "personal"
    if any(k in name for k in config.OVERSEAS):      # 신죠샤·도쿠마 = 해외(분기·JPY)
        return "overseas"
    return vendor_type(name)                          # naver / ridi / kakao / biz


def _load_proof_map(master_path):
    """업체명 → 증빙구분(세금계산서/계산서/혼합)."""
    m = load_workbook(master_path, data_only=True)
    ws = m["업체마스터"]
    proof = {}
    for r in range(4, ws.max_row + 1):
        nm = ws.cell(r, 2).value
        if nm and "예시" not in str(nm):
            proof[str(nm).strip()] = (ws.cell(r, 10).value or config.PROOF_TAXED)
    return proof


def _match(name, mapping):
    if name in mapping:
        return mapping[name]
    for k, v in mapping.items():
        kk = k.split("(")[0].strip()
        if kk and (kk in name or name in k):
            return v
    return None


def _log(rows, path, header):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  → 로그: {path}")


# ── STAGE 1: 정산서 생성 ───────────────────────────────────
def stage_settle(company, y, mo, quarter, groups):
    proof = _load_proof_map(config.MASTER_PATH)
    type_map = _load_type_map(config.MASTER_PATH)
    base = _outbase(company, y, mo)
    outdir = os.path.join(base, "정산서")
    os.makedirs(outdir, exist_ok=True)
    prefix = "수성웹툰" if company == "수성웹툰" else "테라핀"
    srcs = sorted(glob.glob(os.path.join(config.INBOX, f"{prefix}*원작료_정산서_*.xlsx")))
    rows = []
    print(f"[1/3] 정산서 생성 — {company} {y}.{mo:02d}  대상그룹={sorted(groups)}  (원본 {len(srcs)}건)")
    for src in srcs:
        comp, name = _vendor_from_filename(src)
        g = group_of(comp, name)
        if g not in groups:                       # 선택되지 않은 그룹은 건너뜀
            print(f"    · 제외(그룹 미선택: {g}): {name}")
            continue
        vt = date_type(comp, name, type_map)          # 마스터 유형 기준 판정
        if vt in ("ridi", "overseas") and not quarter:
            print(f"    ⚠ 분기 미지정으로 건너뜀(대상 분기 선택 필요): {name}")
            continue
        tax = _match(name, proof) or config.PROOF_TAXED
        out = os.path.join(outdir, f"원작료정산서_{name.replace(' ', '_')}_{y}{mo:02d}.xlsx")
        res = settlement.regen(src, out, name, y, mo, quarter=quarter, tax_type=tax, vtype=vt)
        rows.append([name, tax, res["dates"].get("작성", "-"), res["dates"].get("마감", "-"),
                     res["dates"].get("지급", "-"), res["formula_to_value"],
                     ";".join(res["removed_sheets"]), os.path.basename(out)])
        print(f"    ✓ {name:18} 증빙={tax[:8]:8} 날짜={res['dates']}")
    _log(rows, os.path.join(base, "logs", f"01_정산서_{company}_{y}{mo:02d}.csv"),
         ["업체", "증빙구분", "작성일자", "마감일", "지급일", "수식→값", "제거시트", "파일"])
    print(f"  완료: {len(rows)}건. ▶ 검수 후 'pdf' 단계 진행.")


# ── STAGE 2: PDF + 비밀번호 ────────────────────────────────
def stage_pdf(company, y, mo, groups):
    pwmap = securepdf.build_pwmap(config.MASTER_PATH)
    base = _outbase(company, y, mo)
    indir = os.path.join(base, "정산서")
    pdfdir = os.path.join(base, "PDF")
    tmpdir = os.path.join(base, "_tmp")
    rows = []
    files = sorted(glob.glob(os.path.join(indir, "*.xlsx")))
    print(f"[2/3] PDF + 비밀번호 — 대상그룹={sorted(groups)} (정산서 {len(files)}건)")
    for xlsx in files:
        nm = _name_from_settle_file(xlsx, y, mo)
        if group_of(company, nm) not in groups:       # 선택 그룹만 PDF 변환
            continue
        pw, typ = securepdf.match_pw(xlsx, pwmap)
        if not pw:
            rows.append([os.path.basename(xlsx), "★비밀번호없음", "-"])
            print(f"    ★ 비밀번호 미확보: {os.path.basename(xlsx)}")
            continue
        pdf = securepdf.settlement_only_pdf(xlsx, tmpdir, pdfdir)
        securepdf.encrypt(pdf, pw)
        status = securepdf.verify(pdf, pw)
        rows.append([os.path.basename(pdf), typ, status])
        print(f"    ✓ {os.path.basename(pdf)[:44]:44} [{typ}] {status}")
    _log(rows, os.path.join(base, "logs", f"02_PDF_{company}_{y}{mo:02d}.csv"),
         ["PDF", "유형", "검증"])
    print(f"  완료. ▶ 검수 후 'mail' 단계 진행.")


# ── STAGE 3: 메일 Draft ────────────────────────────────────
def stage_mail(company, y, mo, groups, quarter):
    cfg = config.COMPANIES[company]
    email, nc2id = maildraft.load_email_master(config.MASTER_PATH)
    proof = _load_proof_map(config.MASTER_PATH)
    type_map = _load_type_map(config.MASTER_PATH)
    base = _outbase(company, y, mo)
    pdfdir = os.path.join(base, "PDF")
    outdir = os.path.join(base, "메일Draft")
    os.makedirs(outdir, exist_ok=True)
    rows = []
    pdfs = sorted(glob.glob(os.path.join(pdfdir, "*.pdf")))
    print(f"[3/3] 메일 Draft — {company}  대상그룹={sorted(groups)}")
    for pdf in pdfs:
        comp, name = _vendor_from_filename(pdf.replace(".pdf", ".xlsx"))
        name = re.sub(r"^원작료정산서 ", "", os.path.basename(pdf).replace(".pdf", ""))
        name = re.sub(rf"_{y}{mo:02d}$", "", name.replace("원작료정산서_", "")).replace("_", " ").strip()
        if group_of(company, name) not in groups:     # 선택 그룹만 메일 초안 생성
            continue
        # 개인/사업자 양식키
        is_personal = False
        vid = _match_id(name, nc2id, company)
        key = cfg["mail_key_personal"]
        info = email.get((vid, key)) if (vid and key) else None
        if info:
            is_personal = True
        else:
            info = email.get((vid, cfg["mail_key_biz"])) if vid else None
        if not info or not info["to"] or "example@" in str(info["to"]):
            rows.append([name, "★수신정보없음", "-"])
            print(f"    ★ 수신정보 없음: {name}")
            continue
        vt = vendor_type(name)
        if vt == "ridi":
            d = settle_dates_quarter(y, quarter or (((mo - 1) // 3) + 1))
        else:
            d = settle_dates(vt, y, mo)
        def fmt(x):
            return x.strftime("%Y-%m-%d") if isinstance(x, datetime.datetime) else "-"
        pay = fmt(d.get("지급"))
        if is_personal:
            body = maildraft.body_personal(cfg["display"], pay)
        else:
            body = maildraft.body_biz(cfg["display"], mo, fmt(d.get("작성")),
                                      fmt(d.get("마감")), pay, cfg["send_email"], cfg["biz_license"])
        subj = cfg["subject"].format(y=y, m=mo)
        out = os.path.join(outdir, f"{subj}_{name}.eml")
        maildraft.make_eml(out, info["frm"], maildraft._split(info["to"]),
                           maildraft._split(info["cc"]), maildraft._split(info["bcc"]),
                           subj, body, cfg["signature"], [pdf])
        rows.append([name, "개인" if is_personal else "사업자", "OK"])
        print(f"    ✓ {name:18} To={maildraft._split(info['to'])[0]}")
    _log(rows, os.path.join(base, "logs", f"03_메일_{company}_{y}{mo:02d}.csv"),
         ["업체", "유형", "상태"])
    print("  완료. ▶ Outlook/MailPlug 에서 .eml 열어 검수 후 발송. (사업자등록증 후첨)")


def _match_id(name, nc2id, company):
    주 = "수성웹툰" if company == "수성웹툰" else "테라핀"
    # 그룹 공통(카카오/예원북스) 대표명으로 치환
    for rep, members in config.PW_GROUPS.items():
        if any(_norm(name) == _norm(mb) for mb in members):
            name = rep
            break
    for (n, j), i in nc2id.items():
        if j == 주 and (_norm(n) == _norm(name) or _norm(n.split("(")[0]) == _norm(name)):
            return i
    for (n, j), i in nc2id.items():
        if _norm(n.split("(")[0]) in _norm(name) or _norm(name) in _norm(n):
            return i
    return None


def _norm(s):
    return re.sub(r"[\s_()]", "", str(s)).strip().lower() if s else ""


# ──────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="원작료 정산서 자동화")
    ap.add_argument("--company", default=None, choices=list(config.COMPANIES),
                    help="(구버전 호환) 특정 회사만. 미지정 시 --groups 로 자동 결정")
    ap.add_argument("--groups", default="all",
                    help="실행할 그룹: 'all' 또는 쉼표구분. 예) 테라핀_네이버,테라핀_그외")
    ap.add_argument("--year", type=int, default=config.SETTLE_YEAR)
    ap.add_argument("--month", type=int, default=config.SETTLE_MONTH)
    ap.add_argument("--quarter", type=int, default=config.SETTLE_QUARTER)
    ap.add_argument("--stage", default="all", choices=["settle", "pdf", "mail", "all"])
    a = ap.parse_args()

    # 선택 그룹 해석
    if a.groups == "all":
        sel = set(config.GROUP_KEYS)
    else:
        sel = {g.strip() for g in a.groups.split(",") if g.strip()}
        bad = sel - set(config.GROUP_KEYS)
        if bad:
            sys.exit(f"알 수 없는 그룹: {sorted(bad)}\n사용 가능: {config.GROUP_KEYS}")
    if a.company:                                   # 구버전 호환: 특정 회사로 한정
        keep = {"수성"} if a.company == "수성웹툰" else {g for g in sel if g.startswith("테라핀")}
        sel &= keep
    # 실행할 회사 = 선택 그룹에서 자동 도출
    companies = sorted({"수성웹툰" if g == "수성" else "테라핀" for g in sel})

    print("=" * 60)
    print(f" 원작료 정산서 자동화 | {a.year}.{a.month:02d} | 단계={a.stage}")
    print(f" 대상회사={companies} | 대상그룹={sorted(sel)}")
    print(" 불변원칙: 원본무수정 · 무발송 · 검수게이트")
    print("=" * 60)
    if not sel:
        sys.exit("선택된 그룹이 없습니다.")
    if ({"테라핀_리디", "테라핀_해외"} & sel) and not a.quarter:
        sys.exit("리디·해외 그룹은 --quarter(대상 분기 1~4)를 반드시 지정해야 합니다.\n"
                 "  예) python pipeline.py --groups 테라핀_리디 --quarter 2 --stage settle")

    if a.stage in ("settle", "all"):
        for c in companies:
            stage_settle(c, a.year, a.month, a.quarter, sel)
        if a.stage == "all":
            input("\n[검수 게이트] 정산서를 검수하세요. 계속하려면 Enter…")
    if a.stage in ("pdf", "all"):
        for c in companies:
            stage_pdf(c, a.year, a.month, sel)
        if a.stage == "all":
            input("\n[검수 게이트] PDF·비밀번호를 검수하세요. 계속하려면 Enter…")
    if a.stage in ("mail", "all"):
        for c in companies:
            stage_mail(c, a.year, a.month, sel, a.quarter)
    print("\n끝. 산출물: output/정산서, output/PDF, output/메일Draft, output/logs")


if __name__ == "__main__":
    main()
