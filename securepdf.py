# -*- coding: utf-8 -*-
"""
PDF 변환 + 비밀번호 설정

 - 정산 시트(현재월)만 1페이지 PDF로 변환  (직전월 참고 시트는 PDF에서 제외)
 - 비밀번호: 사업자=사업자등록번호 10자리 / 개인=생년월일 6자리
 - 그룹 공통번호(카카오·KW예원북스)는 마스터/PW_GROUPS로 매핑
 - 무발송 원칙: 암호화 PDF 생성까지만.
"""
import os, re, sys, shutil, subprocess
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

import config

SOFFICE_SKILL = "/mnt/skills/public/xlsx/scripts/office/soffice.py"   # Linux 개발 컨테이너용 래퍼


def _resolve_soffice():
    """LibreOffice 실행 커맨드(앞부분). 우선순위:
       1) SOFFICE_PATH 환경변수  2) Linux 스킬 래퍼  3) Windows 기본 설치 경로  4) PATH."""
    p = os.environ.get("SOFFICE_PATH")
    if p and os.path.exists(p):
        return [sys.executable, p] if p.lower().endswith(".py") else [p]
    if os.path.exists(SOFFICE_SKILL):
        return ["python3", SOFFICE_SKILL]
    for c in (r"C:\Program Files\LibreOffice\program\soffice.exe",
              r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"):
        if os.path.exists(c):
            return [c]
    exe = shutil.which("soffice") or shutil.which("soffice.exe") or shutil.which("soffice.com")
    if exe:
        return [exe]
    raise RuntimeError(
        "LibreOffice(soffice)를 찾지 못했습니다. LibreOffice 설치 후 "
        "SOFFICE_PATH 환경변수로 soffice.exe 경로를 지정하세요.")


def _nmf(s):
    return re.sub(r"[\s_]", "", str(s)).strip().lower() if s else ""


def build_pwmap(master_path):
    """업체명(정규화) → (비밀번호, 유형). 그룹 공통번호 포함."""
    m = load_workbook(master_path, data_only=True)
    ws = m["업체마스터"]
    pwmap = {}
    for r in range(4, ws.max_row + 1):
        nm = ws.cell(r, 2).value
        if not nm or "예시" in str(nm):
            continue
        typ = ws.cell(r, 4).value
        biz = re.sub(r"\D", "", str(ws.cell(r, 7).value or ""))
        per = re.sub(r"\D", "", str(ws.cell(r, 8).value or ""))
        pw = biz if (typ == "사업자" and len(biz) == 10) else (per if (typ == "개인" and len(per) == 6) else None)
        if pw:
            pwmap[_nmf(nm)] = (pw, typ)
            base = re.sub(r"\(.+\)", "", str(nm))   # 괄호 제거 별칭
            pwmap.setdefault(_nmf(base), (pw, typ))
    # 그룹 공통번호
    for rep, members in config.PW_GROUPS.items():
        if _nmf(rep) in pwmap:
            for mb in members:
                pwmap.setdefault(_nmf(mb), pwmap[_nmf(rep)])
    return pwmap


def match_pw(filename, pwmap):
    base = os.path.basename(filename)
    base = re.sub(r"^(원작료정산서_|수성웹툰_원작료정산서_)", "", base).replace("_202605.xlsx", "").replace(".xlsx", "")
    key = _nmf(base)
    if key in pwmap:
        return pwmap[key]
    for k in pwmap:
        if k and (k in key or key in k):
            return pwmap[k]
    return (None, None)


def settlement_only_pdf(xlsx, tmpdir, pdfdir):
    """정산 시트만 남긴 임시 xlsx → PDF."""
    os.makedirs(tmpdir, exist_ok=True)
    os.makedirs(pdfdir, exist_ok=True)
    tmp = os.path.join(tmpdir, os.path.basename(xlsx))
    shutil.copy(xlsx, tmp)
    wb = load_workbook(tmp)
    keep = wb.sheetnames[0]   # settlement.regen 이 정산 시트를 맨 앞으로 둠
    for s in list(wb.sheetnames):
        if s != keep:
            del wb[s]
    wb.save(tmp)
    subprocess.run(_resolve_soffice() + ["--headless", "--convert-to", "pdf",
                    "--outdir", pdfdir, tmp], capture_output=True)
    return os.path.join(pdfdir, os.path.basename(xlsx).replace(".xlsx", ".pdf"))


def encrypt(pdf, pw):
    r = PdfReader(pdf)
    w = PdfWriter()
    for p in r.pages:
        w.add_page(p)
    w.encrypt(pw, pw)   # user = owner = 거래처 비밀번호
    with open(pdf, "wb") as f:
        w.write(f)


def verify(pdf, pw):
    r = PdfReader(pdf)
    if not r.is_encrypted:
        return "NOT_ENCRYPTED"
    return "OK" if r.decrypt(pw) else "WRONG_PW"
