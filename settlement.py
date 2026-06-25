# -*- coding: utf-8 -*-
"""
정산서 생성 — 원본 템플릿 복제 방식 (검증 완료)

생성 절차 (업체별):
  1) 직전 발행 정산서 사본을 복제 (서식 100% 보존)
  2) 수식 → 캐시 계산값으로 치환  ((가)버전: #NAME? 제거, 어느 뷰어에서나 동일)
  3) 작성/마감/지급일을 유형별 규칙으로 기입 (카카오는 '-' 보존)
  4) 모든 날짜 YYYY-MM-DD 강제 (한글 '년/월' 월표시는 보존)
  5) 증빙구분 적용 (면세=계산서·부가세0, 혼합=원본 분리 보존)
  6) 발송 불필요 부가 시트 제거 (KW유통·만타·MG종료·해지 등)
  7) 정산 시트를 맨 앞·활성으로 이동

※ 원본은 INBOX(담당자 첨부 사본)에서만 읽음. NAS 직접 접근 없음.
"""
import os, re, glob, shutil, datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import config
from daterules import settle_dates, settle_dates_quarter, vendor_type


def _pick_sheet(names, y, mo):
    """정산 시트 자동 식별."""
    cands = [f"{mo:02d}", f"{y}-{mo:02d}", f"{y}-{mo:02d}월",
             f"{y}.{((mo-1)//3)+1}Q"]
    for s in cands:
        if s in names:
            return s
    return names[0]


def _force_iso_dates(ws):
    """datetime 셀 → yyyy-mm-dd. 단 한글 '년/월' 월표시 형식은 보존."""
    n = 0
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, datetime.datetime):
                fmt = c.number_format or ""
                if "년" in fmt or "월" in fmt:
                    continue
                if c.number_format != config.RULES["date_format"]:
                    c.number_format = config.RULES["date_format"]
                    n += 1
    return n


def _is_pure_period(s):
    t = s.replace("-", "").replace(".", "").replace("월", "").replace("Q", "").strip()
    return t.isdigit() or bool(re.match(r"^20\d\d", s)) or bool(re.match(r"^\d{1,2}$", s))


def regen(path_in, path_out, vendor_name, y, mo, quarter=None, tax_type="세금계산서", vtype=None):
    """원본 1개 → 정산서 1개 생성. 모든 규칙 적용."""
    shutil.copy(path_in, path_out)
    wb = load_workbook(path_out)
    wbv = load_workbook(path_in, data_only=True)
    sh = _pick_sheet(wb.sheetnames, y, mo)
    ws, wsv = wb[sh], wbv[sh]

    # (2) 수식 → 캐시값
    conv = 0
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                c.value = wsv[c.coordinate].value
                conv += 1

    # (3) 날짜 규칙  (유형은 호출측에서 업체마스터 기준으로 결정해 vtype 으로 전달)
    vt = vtype or vendor_type(vendor_name)
    if vt == "overseas":
        d = dict(작성=None, 마감=None, 지급=None, dash=True)   # 해외 일정 미확정 → 수기 입력
    elif vt == "ridi":
        if not quarter:
            raise ValueError(f"리디 정산({vendor_name})은 대상 분기 지정이 필요합니다.")
        d = settle_dates_quarter(y, quarter)
    else:
        d = settle_dates(vt, y, mo)

    hr, col = None, {}
    for r in range(1, 16):
        for cc in range(1, 18):
            v = ws.cell(r, cc).value
            if isinstance(v, str):
                t = v.strip()
                if t == "작성일자": col["작성"] = cc
                elif t == "마감일": col["마감"] = cc
                elif t == "지급일": col["지급"] = cc
        if col:
            hr = r
            break

    def anchor(r, cc):
        coord = f"{get_column_letter(cc)}{r}"
        for mr in ws.merged_cells.ranges:
            if coord in mr:
                return mr.min_row, mr.min_col
        return r, cc

    if hr and not d.get("dash"):
        rr = hr + 1
        while rr < hr + 40:
            b = ws.cell(rr, 2).value
            if isinstance(b, str) and b.replace(" ", "").strip() == "합계":
                break
            if any(ws.cell(rr, cc).value not in (None, "") for cc in range(2, 8)):
                for key, cc in col.items():
                    if d.get(key) is not None:
                        ar, ac = anchor(rr, cc)
                        if str(ws.cell(ar, ac).value).strip() == "-":  # 카카오 등 '-' 보존
                            continue
                        ws.cell(ar, ac).value = d[key]
            rr += 1

    # (4) 모든 날짜 YYYY-MM-DD
    _force_iso_dates(ws)

    # (5) 증빙구분 — 면세: 계산서 + 부가세0 + 실지급=공급가액  (혼합/과세: 원본 보존)
    if tax_type == config.PROOF_EXEMPT:
        scol, shr = {}, None
        for r in range(1, 16):
            for cc in range(1, 18):
                v = ws.cell(r, cc).value
                if isinstance(v, str):
                    t = v.strip().replace(" ", "")
                    if t == "공급가액": scol["공급"] = cc
                    elif t == "부가세": scol["부가"] = cc
                    elif t in ("실지급액", "실지급"): scol["실지급"] = cc
            if scol.get("공급"):
                shr = r
                break
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip() == config.PROOF_TAXED:
                    c.value = config.PROOF_EXEMPT
        if scol.get("공급") and scol.get("부가"):
            rr = shr + 1
            while rr < shr + 40:
                b = ws.cell(rr, 2).value
                stop = isinstance(b, str) and b.replace(" ", "").strip() == "합계"
                g = ws.cell(rr, scol["공급"]).value
                if isinstance(g, (int, float)):
                    ws.cell(rr, scol["부가"]).value = 0
                    if scol.get("실지급"):
                        ws.cell(rr, scol["실지급"]).value = g
                if stop:
                    break
                rr += 1

    # (6) 부가 시트 제거
    removed = []
    for s in list(wb.sheetnames):
        if not _is_pure_period(s) and any(k in s for k in config.DROP_SHEET_KEYWORDS):
            del wb[s]
            removed.append(s)

    # (7) 정산 시트 맨 앞·활성
    if sh in wb.sheetnames:
        idx = wb.sheetnames.index(sh)
        if idx > 0:
            wb.move_sheet(sh, -idx)
    wb.active = 0
    wb.save(path_out)
    return dict(sheet=sh, formula_to_value=conv,
                dates={k: (v.strftime("%Y-%m-%d") if isinstance(v, datetime.datetime) else "-")
                       for k, v in d.items() if k in ("작성", "마감", "지급")},
                removed_sheets=removed, tax_type=tax_type)
