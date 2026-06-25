# -*- coding: utf-8 -*-
"""
메일 초안(.eml) 생성 — 무발송 원칙 (Draft 까지만)

 - 이메일 마스터에서 발신/수신/참조(CC)/숨은참조(BCC)/양식키를 읽음
 - 본문: 회사·유형별 템플릿 (사업자=계산서 발행요청표 / 개인=지급예정·생년월일 안내)
 - 첨부: 암호화 정산서 PDF (+ 사업자등록증은 프로그램 완성 시 후첨)
 - 카카오 그룹: 수신 공유, 파일/첨부/본문 날짜('-')는 플랫폼별로 개별
 - .eml 은 Outlook/MailPlug 에서 열면 그대로 검토 후 발송 가능
"""
import os, re
from email.message import EmailMessage
from openpyxl import load_workbook

import config


def _clean(s):
    return re.sub(r"[\r\n\t]+", " ", str(s)).strip() if s else ""


def _split(s):
    return [_clean(x) for x in str(s).replace("\n", ";").replace("\r", ";").split(";") if _clean(x)]


def load_email_master(master_path):
    """(업체ID, 양식키) → dict(frm,to,cc,bcc), 그리고 (업체명,정산주체)→ID."""
    m = load_workbook(master_path, data_only=True)
    em, vm = m["이메일마스터"], m["업체마스터"]
    nc2id = {}
    for r in range(4, vm.max_row + 1):
        if vm.cell(r, 1).value and "예시" not in str(vm.cell(r, 2).value or ""):
            nc2id[(str(vm.cell(r, 2).value).strip(), str(vm.cell(r, 3).value or "").strip())] = str(vm.cell(r, 1).value).strip()
    email = {}
    for r in range(4, em.max_row + 1):
        vid = em.cell(r, 1).value
        if not vid or str(vid).startswith("["):
            continue
        email[(str(vid).strip(), str(em.cell(r, 6).value))] = dict(
            frm=em.cell(r, 2).value, to=em.cell(r, 3).value,
            cc=em.cell(r, 4).value, bcc=em.cell(r, 5).value)
    return email, nc2id


def body_personal(company_disp, pay):
    return (f"안녕하세요,\n{company_disp} 정산 담당자입니다.\n\n"
            f"# {pay[:4]}년 {pay[5:7]}월 원작료 정산서 전달 드립니다.\n\n"
            f"1. 파일 첨부 : 원작료 정산서\n"
            f"2. 지급예정일 : {pay}\n"
            f"3. 안내사항 : 정산서 열람 시 비밀번호가 설정되었습니다. (비밀번호 : 생년월일 6자리)\n\n"
            f"   참고하시어 정산서 확인 부탁드리며, 혹 관련하여 오류가 있으실 경우 회신 부탁드리겠습니다.\n\n"
            f"확인하신 후, 문의사항 있으시면 편하게 말씀해주세요.\n감사합니다.\n")


def body_biz(company_disp, mo, asof, due, pay, send_email, biz_license):
    a = asof or "-"
    d = due or "-"
    p = pay or "-"
    return (f"안녕하세요,\n{company_disp} 정산팀입니다.\n\n"
            f"# 2026년 {mo:02d}월 원작료 정산서 전달 드립니다.\n\n"
            f"1. 파일 첨부\n(1) {mo:02d}월 원작료 정산서  (2) {biz_license}\n\n"
            f"2. (세금)계산서 발행 요청   *발행 마감일 : {d}\n"
            f" - 작성일자 : {a}\n"
            f" - (세금)계산서 발송 이메일 : {send_email}\n"
            f" - 품목명 : 원작료_{mo:02d}월 정산 <작품명 기재해주세요.>\n"
            f"   (정산서가 2개 이상일 경우 (세금)계산서는 1장으로 발행해 주시되 품목 나눠 기재 부탁드립니다.)\n"
            f" - 지급일 : {p}\n\n"
            f"3. 안내사항 : 정산서 열람 시 비밀번호가 설정되었습니다. (비밀번호 : 귀사 사업자등록번호)\n"
            f"   참고하시어 정산서 확인 부탁드리며, 혹 관련하여 오류가 있으실 경우 회신 부탁드리겠습니다.\n\n"
            f"확인하신 후, 문의사항 있으시면 편하게 말씀해주세요.\n감사합니다.\n")


def make_eml(out, frm, to, cc, bcc, subject, body, signature, attachments):
    msg = EmailMessage()
    msg["From"] = _clean(frm)
    msg["To"] = ", ".join(to)
    if cc:
        msg["Cc"] = ", ".join(cc)
    if bcc:
        msg["Bcc"] = ", ".join(bcc)
    msg["Subject"] = subject
    msg.set_content(body + "\n" + signature)
    for ap in attachments:
        with open(ap, "rb") as f:
            msg.add_attachment(f.read(), maintype="application", subtype="pdf",
                               filename=os.path.basename(ap))
    with open(out, "wb") as f:
        f.write(bytes(msg))


def _last_day_str(ym):
    """'2026-06' → '2026.06.30'."""
    import calendar
    y, m = int(ym[:4]), int(ym[5:7])
    return f"{y}.{m:02d}.{calendar.monthrange(y, m)[1]:02d}"


def subject_overseas(정산서월):
    """[Terapin] Revenue Share YYYY.MM"""
    return f"[Terapin] Revenue Share {정산서월[:4]}.{정산서월[5:7]}"


def body_overseas(company_en, period_end, amount_jpy, 발행):
    """해외(도쿠마·신죠샤) Revenue Share 메일 본문. 발행=True→Payment, False→Carried Forward."""
    status = "Payment" if 발행 else "Carried Forward"
    amt = f"{amount_jpy:,.2f}" if isinstance(amount_jpy, (int, float)) else str(amount_jpy)
    return (f"Dear {company_en},\n\n"
            f"I hope you are doing well.\n\n"
            f"Please find attached the revenue share report for the period ending {period_end}.\n\n"
            f"Revenue Share Amount:\n"
            f"  * Amount: {amt} JPY\n\n"
            f"Payment Status:\n"
            f"  * {status}\n\n"
            f"Please refer to the attached report for detailed information.\n"
            f"If you have any questions, please feel free to contact us.\n\n"
            f"Best regards,\n")


def overseas_eml(out, email_cfg, company_en, 정산서월, amount_jpy, 발행,
                 attachments, signature=""):
    """해외 정산서 .eml 생성(영문 양식)."""
    subject = subject_overseas(정산서월)
    body = body_overseas(company_en, _last_day_str(정산서월), amount_jpy, 발행)
    to = _split(email_cfg.get("to"))
    cc = _split(email_cfg.get("cc"))
    bcc = _split(email_cfg.get("bcc"))
    make_eml(out, email_cfg.get("frm"), to, cc, bcc, subject, body, signature, attachments)
    return dict(subject=subject, to=to, cc=cc, 발행=발행)
